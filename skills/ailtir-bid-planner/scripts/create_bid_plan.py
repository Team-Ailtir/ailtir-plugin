import argparse
import sys
from datetime import datetime

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
except ImportError:
    print("openpyxl not installed. Please install it.")
    sys.exit(1)

# Ailtir Brand Palette
AILTIR_NAVY = "0A1128" # Navy 900
AILTIR_PURPLE = "7C3AED" # Purple 500
AILTIR_LIGHT = "F5F7FA"
AILTIR_WHITE = "FFFFFF"
FLAG_AMBER = "F59E0B" # Amber 400

HEADER_FONT = Font(bold=True, color=AILTIR_WHITE, size=11, name="Space Grotesk")
HEADER_FILL = PatternFill("solid", fgColor=AILTIR_NAVY)
BODY_FONT = Font(size=10, name="Inter")
THIN_BORDER = Border(
    left=Side(style="thin", color="D9D9D9"),
    right=Side(style="thin", color="D9D9D9"),
    top=Side(style="thin", color="D9D9D9"),
    bottom=Side(style="thin", color="D9D9D9"),
)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT_WRAP = Alignment(horizontal="left", vertical="top", wrap_text=True)

def style_header(ws, row, max_col):
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        cell.border = THIN_BORDER

def build_summary(wb, project, client, return_date, route):
    ws = wb.active
    ws.title = "1. Bid Summary"
    ws.merge_cells("A1:H1")
    ws["A1"] = f"AILTIR BID PLAN — {project.upper()}"
    ws["A1"].font = Font(bold=True, color=AILTIR_WHITE, size=14, name="Space Grotesk")
    ws["A1"].fill = PatternFill("solid", fgColor=AILTIR_NAVY)
    ws["A1"].alignment = CENTER
    
    details = [
        ("Project Name:", project),
        ("Client:", client),
        ("Tender Return:", return_date),
        ("Procurement Route:", route)
    ]
    
    row = 3
    for label, val in details:
        ws.cell(row=row, column=1, value=label).font = Font(bold=True, name="Inter")
        ws.cell(row=row, column=2, value=val).font = BODY_FONT
        row += 1

def build_register(wb):
    ws = wb.create_sheet("2. Document Register")
    headers = ["Filename", "Title", "Type", "Rev", "Date", "Notes"]
    for i, h in enumerate(headers, 1):
        ws.cell(row=1, column=i, value=h)
    style_header(ws, 1, len(headers))

def build_gonogo(wb):
    ws = wb.create_sheet("3. Go-No-Go")
    headers = ["Criteria", "Max Score", "Actual Score", "Notes"]
    for i, h in enumerate(headers, 1):
        ws.cell(row=1, column=i, value=h)
    style_header(ws, 1, len(headers))

def build_compliance(wb):
    ws = wb.create_sheet("4. Compliance Matrix")
    headers = ["Ref", "Requirement", "Format", "Template Provided", "Owner", "Status"]
    for i, h in enumerate(headers, 1):
        ws.cell(row=1, column=i, value=h)
    style_header(ws, 1, len(headers))

def build_risk(wb):
    ws = wb.create_sheet("5. Risk Register")
    headers = ["Risk ID", "Category", "Description", "Contract Clause", "Mitigation", "Status"]
    for i, h in enumerate(headers, 1):
        ws.cell(row=1, column=i, value=h)
    style_header(ws, 1, len(headers))

def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--output", required=True)
    parser.add_argument("--project", required=True)
    parser.add_argument("--client", default="TBC")
    parser.add_argument("--return-date", default="TBC")
    parser.add_argument("--route", default="TBC")
    args = parser.parse_args()

    wb = Workbook()
    build_summary(wb, args.project, args.client, args.return_date, args.route)
    build_register(wb)
    build_gonogo(wb)
    build_compliance(wb)
    build_risk(wb)
    
    wb.save(args.output)
    print(f"Created {args.output}")

if __name__ == "__main__":
    main()
