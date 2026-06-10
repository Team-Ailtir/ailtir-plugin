#!/usr/bin/env python3
"""
Recalculate Excel formulas after programmatic generation.
openpyxl doesn't calculate formulas - this forces recalc on open.
"""

import sys
from pathlib import Path

def set_recalc_flag(filepath: str) -> None:
    """Set calcMode to auto so Excel recalculates on open."""
    try:
        from openpyxl import load_workbook
    except ImportError:
        print("Installing openpyxl...")
        import subprocess
        subprocess.run([sys.executable, "-m", "pip", "install", "openpyxl", "--break-system-packages"], check=True)
        from openpyxl import load_workbook
    
    wb = load_workbook(filepath)
    
    # Force recalculation on open
    if wb.calculation is None:
        from openpyxl.workbook.properties import CalcProperties
        wb.calculation = CalcProperties()
    wb.calculation.calcMode = "auto"
    wb.calculation.fullCalcOnLoad = True
    
    wb.save(filepath)
    print(f"✓ Recalc flag set: {filepath}")

if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Usage: python recalc.py <excel_file.xlsx>")
        sys.exit(1)
    
    for f in sys.argv[1:]:
        if Path(f).exists():
            set_recalc_flag(f)
        else:
            print(f"✗ File not found: {f}")
