import argparse
import sys

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
except ImportError:
    print("openpyxl not installed. Please install it.")
    sys.exit(1)

AILTIR_NAVY = "0A1128" # Navy 900
AILTIR_WHITE = "FFFFFF"
HEADER_FONT = Font(bold=True, color=AILTIR_WHITE, size=11, name="Space Grotesk")
HEADER_FILL = PatternFill("solid", fgColor=AILTIR_NAVY)
THIN_BORDER = Border(
    left=Side(style="thin", color="D9D9D9"),
    right=Side(style="thin", color="D9D9D9"),
    top=Side(style="thin", color="D9D9D9"),
    bottom=Side(style="thin", color="D9D9D9"),
)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)

def style_header(ws, row, max_col):
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        cell.border = THIN_BORDER

def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--output", required=True)
    parser.add_argument("--package", required=True)
    args = parser.parse_args()

    wb = Workbook()
    
    # Summary Tab
    ws1 = wb.active
    ws1.title = "1. Executive Summary"
    ws1.merge_cells("A1:E1")
    ws1["A1"] = f"QUOTE COMPARISON — {args.package.upper()}"
    ws1["A1"].font = Font(bold=True, color=AILTIR_WHITE, size=14, name="Space Grotesk")
    ws1["A1"].fill = PatternFill("solid", fgColor=AILTIR_NAVY)
    ws1["A1"].alignment = CENTER
    
    headers1 = ["Subcontractor", "Total Quoted", "Variance from Lowest", "Scope Coverage", "Notes"]
    for i, h in enumerate(headers1, 1):
        ws1.cell(row=3, column=i, value=h)
    style_header(ws1, 3, len(headers1))

    # Leveling Tab
    ws2 = wb.create_sheet("2. Normalisation")
    headers2 = ["Scope Item", "Sub A", "Sub B", "Sub C"]
    for i, h in enumerate(headers2, 1):
        ws2.cell(row=1, column=i, value=h)
    style_header(ws2, 1, len(headers2))

    wb.save(args.output)
    print(f"Created {args.output}")

if __name__ == "__main__":
    main()
