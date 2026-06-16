#!/usr/bin/env python3
"""
Excel Register Output for Construction Takeoffs

Generates a 2-sheet Excel workbook:
1. Data — Hierarchical BoQ grouped by trade (Trade / Primary / Secondary rows)
2. Flags — Everything that needs human attention in one list

Can be called from Python or CLI:
    python excel_output.py --json /tmp/final_boq.json --out register.xlsx
"""

import sys
import json
import argparse

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("ERROR: openpyxl required. Install with: pip install openpyxl", file=sys.stderr)
    sys.exit(1)


# ── Style Constants ──

HEADER_FONT = Font(bold=True, color="FFFFFF", name="Arial", size=10)
HEADER_FILL = PatternFill("solid", fgColor="1F497D")
HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)

THIN_BORDER = Border(
    left=Side(style="thin", color="D5D8DC"),
    right=Side(style="thin", color="D5D8DC"),
    top=Side(style="thin", color="D5D8DC"),
    bottom=Side(style="thin", color="D5D8DC"),
)

# Trade subheader row styling
TRADE_FILL = PatternFill("solid", fgColor="D9E1F2")
TRADE_FONT = Font(bold=True, size=13, color="1F497D", name="Arial")

# Primary item row styling
PRIMARY_FONT = Font(bold=True, size=11, name="Arial")

# Secondary / derived item row styling
SECONDARY_FONT = Font(italic=True, size=11, color="595959", name="Arial")
SECONDARY_INDENT = Alignment(indent=3, vertical="center", wrap_text=True)

# Confidence colour coding
CONFIDENCE_FONTS = {
    "HIGH":   Font(color="1B7A1B", name="Arial", size=10, bold=True),
    "MEDIUM": Font(color="CC7A00", name="Arial", size=10),
    "LOW":    Font(color="CC0000", name="Arial", size=10, bold=True),
}

# Flag severity styling
FLAG_FILLS = {
    "LOW_CONFIDENCE":      PatternFill("solid", fgColor="F5D7D7"),
    "VALIDATION_WARNING":  PatternFill("solid", fgColor="FFF3CD"),
    "DISCREPANCY":         PatternFill("solid", fgColor="F5D7D7"),
    "REVIEW_REQUIRED":     PatternFill("solid", fgColor="FFF3CD"),
    "BENCHMARK_OUTLIER":   PatternFill("solid", fgColor="FFF3CD"),
    "SANITY_CHECK":        PatternFill("solid", fgColor="F5D7D7"),
}


def _apply_header_row(ws, headers, row=1):
    """Apply formatted header row with freeze and filter."""
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN
        cell.border = THIN_BORDER
    ws.freeze_panes = f"A{row + 1}"
    ws.auto_filter.ref = f"A{row}:{get_column_letter(len(headers))}{row}"


def _auto_fit_columns(ws, min_width=10, max_width=50):
    """Auto-fit column widths based on content."""
    for col in range(1, ws.max_column + 1):
        max_length = 0
        for row in range(1, min(ws.max_row + 1, 200)):
            cell = ws.cell(row=row, column=col)
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[get_column_letter(col)].width = min(max(max_length + 2, min_width), max_width)


def _write_row(ws, row_num, values):
    """Write a row with borders."""
    for col, value in enumerate(values, 1):
        cell = ws.cell(row=row_num, column=col, value=value)
        cell.border = THIN_BORDER
        cell.alignment = Alignment(vertical="center", wrap_text=True)


# ── Sheet 1: Data ──

def build_data_sheet(wb, quantities):
    """Hierarchical BoQ grouped by trade.

    Row types:
    - Trade subheader: merged, coloured, large font
    - Primary item: bold
    - Secondary item: italic, indented, grey
    """
    ws = wb.active
    ws.title = "Data"

    headers = [
        "Trade", "Tag", "Description", "Quantity", "Unit",
        "Drawing / Page", "Confidence", "Method", "Assumptions"
    ]
    _apply_header_row(ws, headers)

    # Group quantities by trade
    from collections import OrderedDict
    trades = OrderedDict()
    for q in quantities:
        trade = q.get("trade", "unknown")
        if trade not in trades:
            trades[trade] = []
        trades[trade].append(q)

    row = 2
    for trade, items in trades.items():

        # ── Trade subheader row ──
        _write_row(ws, row, [trade.upper(), "", "", "", "", "", "", "", ""])
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=row, column=col)
            cell.fill = TRADE_FILL
            cell.font = TRADE_FONT
        row += 1

        # ── Item rows ──
        for q in items:
            is_secondary = q.get("item_type") == "secondary"

            values = [
                "",  # Trade column blank for items (shown in subheader)
                q.get("tag", ""),
                q.get("description", ""),
                q.get("quantity"),
                q.get("unit", ""),
                q.get("drawing_ref", ""),
                q.get("confidence", ""),
                q.get("method", ""),
                q.get("note", q.get("assumptions", "")),
            ]
            _write_row(ws, row, values)

            if is_secondary:
                # Indented, italic, grey
                for col in range(1, len(headers) + 1):
                    ws.cell(row=row, column=col).font = SECONDARY_FONT
                ws.cell(row=row, column=3).alignment = SECONDARY_INDENT
            else:
                # Bold primary item
                for col in range(1, len(headers) + 1):
                    ws.cell(row=row, column=col).font = PRIMARY_FONT

            # Confidence colour
            conf = q.get("confidence", "")
            if conf in CONFIDENCE_FONTS:
                ws.cell(row=row, column=7).font = CONFIDENCE_FONTS[conf]

            row += 1

    # ── Set column widths ──
    ws.column_dimensions['A'].width = 16   # Trade
    ws.column_dimensions['B'].width = 16   # Tag
    ws.column_dimensions['C'].width = 50   # Description
    ws.column_dimensions['D'].width = 12   # Quantity
    ws.column_dimensions['E'].width = 10   # Unit
    ws.column_dimensions['F'].width = 18   # Drawing/Page
    ws.column_dimensions['G'].width = 14   # Confidence
    ws.column_dimensions['H'].width = 20   # Method
    ws.column_dimensions['I'].width = 45   # Assumptions


# ── Sheet 2: Flags ──

def build_flags_sheet(wb, quantities, validation_results):
    """Everything that needs human attention in one list."""
    ws = wb.create_sheet("Flags")

    headers = [
        "Item", "Tag", "Trade", "Quantity", "Unit",
        "Confidence", "Flag Type", "Details"
    ]
    _apply_header_row(ws, headers)

    row = 2
    flags_written = 0

    # 1. Low confidence items from quantities
    for q in quantities:
        if q.get("confidence") == "LOW":
            values = [
                q.get("description", ""),
                q.get("tag", ""),
                q.get("trade", ""),
                q.get("quantity"),
                q.get("unit", ""),
                "LOW",
                "LOW_CONFIDENCE",
                q.get("note", q.get("assumptions", "Requires manual verification")),
            ]
            _write_row(ws, row, values)
            ws.cell(row=row, column=6).font = CONFIDENCE_FONTS.get("LOW", Font())
            ws.cell(row=row, column=7).fill = FLAG_FILLS.get("LOW_CONFIDENCE", PatternFill())
            row += 1
            flags_written += 1

    # 2. Items explicitly flagged for review
    review_data = validation_results.get("review", {})
    for item in review_data.get("review_items", []):
        values = [
            item.get("item", item.get("description", "")),
            item.get("tag", ""),
            item.get("trade", ""),
            item.get("quantity"),
            item.get("unit", ""),
            item.get("confidence", ""),
            "REVIEW_REQUIRED",
            "; ".join(item.get("review_reasons", [])),
        ]
        _write_row(ws, row, values)
        ws.cell(row=row, column=7).fill = FLAG_FILLS.get("REVIEW_REQUIRED", PatternFill())
        row += 1
        flags_written += 1

    # 3. Validation warnings and errors
    for layer_name, layer_results in validation_results.get("layers", {}).items():
        for result in layer_results:
            status = result.get("status", "")
            if status in ("warning", "error"):
                flag_type = "SANITY_CHECK" if "sanity" in layer_name.lower() else "VALIDATION_WARNING"
                if "benchmark" in layer_name.lower():
                    flag_type = "BENCHMARK_OUTLIER"

                values = [
                    result.get("check", "").replace("_", " ").title(),
                    "",
                    "",
                    result.get("actual", ""),
                    "",
                    "",
                    flag_type,
                    result.get("message", ""),
                ]
                _write_row(ws, row, values)
                ws.cell(row=row, column=7).fill = FLAG_FILLS.get(flag_type, PatternFill())
                row += 1
                flags_written += 1

    # 4. Sanity check failures
    for result in validation_results.get("sanity_checks", []):
        status = result.get("status", "")
        if status in ("warning", "error"):
            values = [
                result.get("check", "").replace("_", " ").title(),
                "",
                "",
                "",
                "",
                "",
                "SANITY_CHECK",
                result.get("message", ""),
            ]
            _write_row(ws, row, values)
            ws.cell(row=row, column=7).fill = FLAG_FILLS.get("SANITY_CHECK", PatternFill())
            row += 1
            flags_written += 1

    # If no flags, write a single "all clear" row
    if flags_written == 0:
        ws.cell(row=2, column=1, value="No items flagged for review.").font = Font(
            italic=True, color="1B7A1B", name="Arial", size=11
        )

    _auto_fit_columns(ws)


# ── Main Generator ──

def generate_takeoff_register(quantities, validation_results=None, page_extractions=None,
                              output_path="takeoff_register.xlsx", project_params=None):
    """Generate the 2-sheet Excel workbook (Data + Flags).

    Args:
        quantities: list of quantity dicts (from assembly engine, derive_quantities, or manual)
        validation_results: dict from validate.run_full_validation() (optional)
        page_extractions: list of per-page extraction results (unused — kept for API compat)
        output_path: where to save the .xlsx file
        project_params: dict of project parameters (unused — kept for API compat)

    Returns:
        str: path to the generated file
    """
    if validation_results is None:
        validation_results = {"layers": {}, "sanity_checks": [], "review": {}, "summary": {}}

    wb = openpyxl.Workbook()

    build_data_sheet(wb, quantities)
    build_flags_sheet(wb, quantities, validation_results)

    output_path = str(output_path)
    wb.save(output_path)
    print(f"Takeoff register saved to: {output_path}", file=sys.stderr)

    return output_path


# ── CLI ──

def main():
    parser = argparse.ArgumentParser(description="Generate 2-sheet takeoff Excel register")
    parser.add_argument("--json", required=True, help="Path to quantities JSON file")
    parser.add_argument("--out", required=True, help="Output Excel path")
    parser.add_argument("--validation", help="Optional validation results JSON")

    args = parser.parse_args()

    with open(args.json) as f:
        data = json.load(f)

    # Support both flat list and wrapped format
    if isinstance(data, list):
        quantities = data
    else:
        quantities = data.get("quantities", data.get("all_quantities", []))

    validation_results = None
    if args.validation:
        with open(args.validation) as f:
            validation_results = json.load(f)

    generate_takeoff_register(
        quantities=quantities,
        validation_results=validation_results,
        output_path=args.out,
    )


if __name__ == "__main__":
    main()
