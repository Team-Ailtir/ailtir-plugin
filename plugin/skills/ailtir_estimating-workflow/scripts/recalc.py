"""Force Excel to recalculate every formula on open.

openpyxl writes stale cached results; setting <calcPr calcMode="auto"
fullCalcOnLoad="true"/> in workbook.xml tells Excel to recompute on open.

Refs: openpyxl CalcProperties (openpyxl.readthedocs.io/en/stable/api/
openpyxl.workbook.properties.html); ECMA-376 Part 1 Section 18.2.2.
"""
from __future__ import annotations

import argparse
import sys
from pathlib import Path

try:
    from openpyxl import load_workbook
    from openpyxl.workbook.properties import CalcProperties
except ImportError:
    sys.stderr.write("ERROR: openpyxl required. Install: "
                     "pip install openpyxl --break-system-packages\n")
    sys.exit(3)

PROFILE_CURRENCY = {"ireland-gc": "€", "uk-gc": "£"}


def _process(path: Path, profile: str | None, quiet: bool) -> bool:
    wb = load_workbook(path)
    if wb.calculation is None:
        wb.calculation = CalcProperties()
    wb.calculation.calcMode = "auto"
    wb.calculation.fullCalcOnLoad = True

    mismatch = False
    if profile:
        want = PROFILE_CURRENCY[profile]
        other = PROFILE_CURRENCY["uk-gc" if profile == "ireland-gc" else "ireland-gc"]
        for ws in wb.worksheets:
            for row in ws.iter_rows():
                for cell in row:
                    fmt = cell.number_format or ""
                    if other in fmt and want not in fmt:
                        mismatch = True
                        sys.stderr.write(
                            f"[WARN] {path.name}: {ws.title}!{cell.coordinate} "
                            f"'{fmt}' - expected {want} for {profile}\n")
    wb.save(path)
    if not quiet:
        print(f"[OK] {path.name}")
    return mismatch


def main(argv: list[str] | None = None) -> int:
    ap = argparse.ArgumentParser(description="Force full recalc on next Excel open.")
    ap.add_argument("files", nargs="+", help="One or more .xlsx workbooks.")
    ap.add_argument("--profile", choices=sorted(PROFILE_CURRENCY), default=None)
    ap.add_argument("--quiet", action="store_true")
    args = ap.parse_args(argv)

    bad_input = any_err = any_mismatch = False
    for raw in args.files:
        p = Path(raw)
        if not p.exists() or p.suffix.lower() != ".xlsx":
            sys.stderr.write(f"[SKIP] {raw}: missing or not .xlsx\n")
            bad_input = True
            continue
        try:
            any_mismatch |= _process(p, args.profile, args.quiet)
        except Exception as exc:
            sys.stderr.write(f"[FAIL] {p.name}: {exc}\n")
            any_err = True

    if bad_input:
        return 2
    if any_mismatch:
        return 4
    return 1 if any_err else 0


if __name__ == "__main__":
    raise SystemExit(main())
