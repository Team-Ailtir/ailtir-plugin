"""style_excel.py — Ailtir brand styling for estimating workbooks.

Palette, typography, borders and profile-aware currency formats applied
to a workbook produced upstream. Styling only: no cell values, formulas,
or structure are added or repaired. Refs: openpyxl.styles, ECMA-376
OOXML colour/font attributes, Ailtir brand system.

Exit codes: 0 ok, 2 missing file, 3 openpyxl unavailable,
4 invalid profile, 5 invalid column letter / row number.
"""

from __future__ import annotations

import argparse
import os
import re
import sys
from typing import Iterable, Optional, Sequence

try:
    from openpyxl import load_workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import column_index_from_string, get_column_letter
    from openpyxl.worksheet.worksheet import Worksheet
except ImportError:
    sys.stderr.write("openpyxl is required. Install with: pip install openpyxl\n")
    sys.exit(3)


# --- Ailtir brand palette (hex, no leading '#') ---------------------------
NAVY_900 = "0A1128"      # Ailtir Navy 900   — primary header fill
NAVY_700 = "1A2550"      # Ailtir Navy 700   — section-row fill
PURPLE_600 = "6D28D9"    # Ailtir Purple 600 — subtotal-row fill
PURPLE_500 = "7C3AED"    # Ailtir Purple 500 — highlight accent
AMBER_400 = "F59E0B"     # Ailtir Amber 400  — warning/flag
LIGHT_ALT = "F5F7FA"     # Ailtir Light      — alternating body row
WHITE = "FFFFFF"         # White             — text on dark fills
TEXT_PRIMARY = "0A1128"  # Ailtir Text prim  — body text on light fills
BORDER_GREY = "D9D9D9"   # Ailtir soft border— softer than default black

# --- Typography (Ailtir brand system) & profile-aware currency ------------
HEADING_FONT_FAMILY = "Space Grotesk"
BODY_FONT_FAMILY = "Inter"
HEADING_PT = 11
BODY_PT = 10
CURRENCY_FORMATS = {
    "ireland-gc": "€#,##0.00;[Red]-€#,##0.00",
    "uk-gc": "£#,##0.00;[Red]-£#,##0.00",
}

# --- Auto-detection vocabularies ------------------------------------------
CURRENCY_HEADER_TOKENS = (
    "rate", "cost", "total", "value", "sum", "price", "amount", "€", "£",
    "unit rate", "subtotal", "net", "gross", "provisional sum", "contingency",
)
SUBTOTAL_PREFIXES = ("subtotal", "sub-total", "trade total",
                     "section total", "element total")
SECTION_PREFIXES = ("element", "section", "group", "trade")


# --- Style primitives -----------------------------------------------------
def _soft_border() -> Border:
    s = Side(border_style="thin", color=BORDER_GREY)
    return Border(left=s, right=s, top=s, bottom=s)

def _fill(hex_rgb: str) -> PatternFill:
    return PatternFill(fill_type="solid", start_color=hex_rgb, end_color=hex_rgb)

def _heading_font() -> Font:
    return Font(name=HEADING_FONT_FAMILY, size=HEADING_PT, bold=True, color=WHITE)

def _body_font() -> Font:
    return Font(name=BODY_FONT_FAMILY, size=BODY_PT, color=TEXT_PRIMARY)

def _apply_row(ws: Worksheet, row: int, hex_rgb: str, align: Alignment) -> None:
    fill, font, border = _fill(hex_rgb), _heading_font(), _soft_border()
    for col in range(1, (ws.max_column or 1) + 1):
        c = ws.cell(row=row, column=col)
        c.fill, c.font, c.alignment, c.border = fill, font, align, border


# --- Public helpers (importable by other estimating scripts) --------------
def style_header_row(ws: Worksheet, row: int = 1) -> None:
    """Navy-900 fill, white Space Grotesk 11pt bold, center/center/wrap."""
    _apply_row(ws, row, NAVY_900,
               Alignment(horizontal="center", vertical="center", wrap_text=True))

def style_section_row(ws: Worksheet, row: int) -> None:
    """Navy-700 fill, white Space Grotesk 11pt bold, left/center/wrap."""
    _apply_row(ws, row, NAVY_700,
               Alignment(horizontal="left", vertical="center", wrap_text=True))

def style_subtotal_row(ws: Worksheet, row: int) -> None:
    """Purple-600 fill (Ailtir accent), white Space Grotesk 11pt bold."""
    _apply_row(ws, row, PURPLE_600,
               Alignment(horizontal="left", vertical="center", wrap_text=True))


def format_currency_column(ws: Worksheet, col_letter: str, profile_key: str) -> None:
    """Apply profile-appropriate currency number format down a column,
    skipping the header row. Sets right/center alignment."""
    fmt = CURRENCY_FORMATS.get(profile_key)
    if fmt is None:
        return
    col_idx = column_index_from_string(col_letter)
    align = Alignment(horizontal="right", vertical="center")
    for row in range(2, (ws.max_row or 1) + 1):
        c = ws.cell(row=row, column=col_idx)
        c.number_format = fmt
        c.alignment = align


def auto_column_width(ws: Worksheet, min_width: int = 10, max_width: int = 60) -> None:
    """Auto-fit column widths. max_width=60 fits NRM2 descriptions such as
    'In situ concrete grade C32/40 in reinforced ground beams over 300mm thick'."""
    max_col = ws.max_column or 1
    max_row = ws.max_row or 1
    for col_idx in range(1, max_col + 1):
        widest = min_width
        for row in range(1, max_row + 1):
            v = ws.cell(row=row, column=col_idx).value
            if v is None:
                continue
            for line in str(v).splitlines() or [str(v)]:
                widest = max(widest, len(line) + 2)
        ws.column_dimensions[get_column_letter(col_idx)].width = min(widest, max_width)


def apply_body_styling(ws: Worksheet, start_row: int = 2,
                       end_row: Optional[int] = None,
                       alt_rows: bool = True) -> None:
    """Inter 10pt body font, soft-grey thin borders, optional alt fill."""
    if end_row is None:
        end_row = ws.max_row or start_row
    if end_row < start_row:
        return
    font, border = _body_font(), _soft_border()
    align = Alignment(horizontal="left", vertical="top", wrap_text=True)
    alt_fill = _fill(LIGHT_ALT)
    last_col = ws.max_column or 1
    for row in range(start_row, end_row + 1):
        shade = alt_rows and ((row - start_row) % 2 == 1)
        for col in range(1, last_col + 1):
            c = ws.cell(row=row, column=col)
            c.font, c.alignment, c.border = font, align, border
            if shade:
                c.fill = alt_fill


# --- Auto-detection -------------------------------------------------------
def _autodetect_currency_cols(ws: Worksheet, header_row: int) -> list[str]:
    hits: list[str] = []
    for col in range(1, (ws.max_column or 1) + 1):
        v = ws.cell(row=header_row, column=col).value
        if v is None:
            continue
        needle = str(v).strip().lower()
        if any(tok in needle for tok in CURRENCY_HEADER_TOKENS):
            hits.append(get_column_letter(col))
    return hits


def _autodetect_row_kinds(ws: Worksheet) -> tuple[list[int], list[int]]:
    """Infer (subtotal_rows, section_rows) from column A prefixes and caps."""
    subs, secs = [], []
    for row in range(2, (ws.max_row or 1) + 1):
        v = ws.cell(row=row, column=1).value
        if v is None:
            continue
        raw = str(v).strip()
        low = raw.lower()
        if any(low.startswith(p) for p in SUBTOTAL_PREFIXES):
            subs.append(row)
        elif any(low.startswith(p) for p in SECTION_PREFIXES):
            secs.append(row)
        else:
            letters = re.sub(r"[^A-Za-z]", "", raw)
            if len(raw) > 8 and letters and raw.upper() == raw:
                secs.append(row)
    return subs, secs


# --- Orchestrator ---------------------------------------------------------
def style_workbook(filepath: str, profile_key: str,
                   currency_cols: Optional[Sequence[str]] = None,
                   subtotal_rows: Optional[Sequence[int]] = None,
                   section_rows: Optional[Sequence[int]] = None,
                   header_rows: Iterable[int] = (1,),
                   alt_rows: bool = True,
                   autodetect_rows: bool = True,
                   quiet: bool = False) -> None:
    """End-to-end styling pass; mutates and saves the workbook in place."""
    wb = load_workbook(filepath)
    header_rows = tuple(header_rows) or (1,)
    for ws in wb.worksheets:
        for hr in header_rows:
            style_header_row(ws, row=hr)
        cols = (list(currency_cols) if currency_cols is not None
                else _autodetect_currency_cols(ws, header_rows[0]))
        if not cols and currency_cols is None and not quiet:
            sys.stderr.write(f"[style_excel] no currency columns detected on '{ws.title}'\n")
        auto_sub, auto_sec = ([], [])
        if autodetect_rows and (subtotal_rows is None or section_rows is None):
            auto_sub, auto_sec = _autodetect_row_kinds(ws)
        subs = list(subtotal_rows) if subtotal_rows is not None else auto_sub
        secs = list(section_rows) if section_rows is not None else auto_sec
        apply_body_styling(ws, start_row=max(header_rows) + 1,
                           end_row=ws.max_row or 1, alt_rows=alt_rows)
        for row in secs:
            style_section_row(ws, row)
        for row in subs:
            style_subtotal_row(ws, row)
        for letter in cols:
            format_currency_column(ws, letter, profile_key)
        auto_column_width(ws)
    wb.save(filepath)


# --- CLI ------------------------------------------------------------------
def _parse_int_list(raw: str, flag: str) -> list[int]:
    out: list[int] = []
    for chunk in filter(None, (c.strip() for c in raw.split(","))):
        try:
            n = int(chunk)
        except ValueError:
            sys.stderr.write(f"invalid integer in {flag}: {chunk!r}\n"); sys.exit(5)
        if n < 1:
            sys.stderr.write(f"row numbers in {flag} must be >= 1\n"); sys.exit(5)
        out.append(n)
    return out


def _parse_col_list(raw: str) -> list[str]:
    out: list[str] = []
    for letter in filter(None, (c.strip().upper() for c in raw.split(","))):
        if not re.fullmatch(r"[A-Z]+", letter):
            sys.stderr.write(f"invalid column letter: {letter!r}\n"); sys.exit(5)
        out.append(letter)
    return out


def _build_parser() -> argparse.ArgumentParser:
    p = argparse.ArgumentParser(description="Apply Ailtir brand styling to an estimating workbook.")
    p.add_argument("workbook", help="path to .xlsx to style in place")
    p.add_argument("--profile", required=True, choices=sorted(CURRENCY_FORMATS.keys()),
                   help="profile key controlling currency symbol")
    p.add_argument("--currency-cols", default=None, help="comma-separated column letters (auto-detected if omitted)")
    p.add_argument("--subtotal-rows", default=None, help="comma-separated 1-indexed subtotal row numbers")
    p.add_argument("--section-rows", default=None, help="comma-separated 1-indexed section row numbers")
    p.add_argument("--header-rows", default="1", help="comma-separated 1-indexed header row numbers (default 1)")
    p.add_argument("--no-alt-rows", action="store_true", help="disable alternating body row shading")
    p.add_argument("--no-autodetect-rows", action="store_true", help="disable row auto-detection")
    p.add_argument("--quiet", action="store_true", help="suppress success stdout")
    return p


def main(argv: Optional[Sequence[str]] = None) -> int:
    a = _build_parser().parse_args(argv)
    if not os.path.isfile(a.workbook):
        sys.stderr.write(f"workbook not found: {a.workbook}\n"); return 2
    if a.profile not in CURRENCY_FORMATS:
        sys.stderr.write(f"unsupported profile: {a.profile}\n"); return 4
    cc = _parse_col_list(a.currency_cols) if a.currency_cols else None
    sr = _parse_int_list(a.subtotal_rows, "--subtotal-rows") if a.subtotal_rows else None
    sc = _parse_int_list(a.section_rows, "--section-rows") if a.section_rows else None
    hr = _parse_int_list(a.header_rows, "--header-rows") or [1]
    style_workbook(a.workbook, a.profile, currency_cols=cc,
                   subtotal_rows=sr, section_rows=sc, header_rows=hr,
                   alt_rows=not a.no_alt_rows,
                   autodetect_rows=not a.no_autodetect_rows, quiet=a.quiet)
    if not a.quiet:
        sys.stdout.write(f"styled: {a.workbook} [{a.profile}]\n")
    return 0


if __name__ == "__main__":
    sys.exit(main())
