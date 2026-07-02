#!/usr/bin/env python3
"""
Ailtir Estimating Workflow — Main Estimate Workbook Generator

Produces a 6-sheet Excel workbook following NRM2 elemental structure. The
workbook is currency- and profile-aware — pass --profile-key to select the
sample prelims rates and currency symbol:

  --profile-key ireland-gc  →  € labour @ SEO 2025 (default; backwards-compatible)
  --profile-key uk-gc       →  £ labour @ CIJC 2026

Usage: python create_estimate.py --project "School Extension" --value 5000000 \
           --profile-key uk-gc --output estimate.xlsx
"""
import sys
import json
import argparse
from pathlib import Path
from datetime import date

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    import subprocess
    subprocess.run([sys.executable, "-m", "pip", "install", "openpyxl", "--break-system-packages", "-q"])
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

# ── Ailtir Brand Colours (v2.6 — DO NOT CHANGE) ──
AILTIR_NAVY    = "0A1128"   # Navy 900 — primary header background
AILTIR_NAVY700 = "1A2550"   # Navy 700 — section divider rows
AILTIR_PURPLE  = "6D28D9"   # Purple 600 — total/summary rows
AILTIR_AMBER   = "F59E0B"   # Amber 400 — key cost figures, urgency
AILTIR_LIGHT   = "F5F7FA"   # Off-white — alternating row fill
WHITE          = "FFFFFF"
BORDER_COLOUR  = "E2E8F0"   # Standard cell border

THIN_BORDER = Border(
    left=Side(style='thin', color=BORDER_COLOUR),
    right=Side(style='thin', color=BORDER_COLOUR),
    top=Side(style='thin', color=BORDER_COLOUR),
    bottom=Side(style='thin', color=BORDER_COLOUR)
)

# ── Profile-driven currency and sample-rate defaults ──
# The script's per-cell number_format strings and sample prelims rates are
# populated from this table at runtime so the same workbook generator can
# emit an Irish estimate or a UK estimate.
PROFILE_DEFAULTS = {
    "ireland-gc": {
        "currency_symbol": "€",
        "number_format": '€#,##0.00',
        "prelims_samples": [
            # (item_no, description, unit, qty, rate)
            ("1.1", "Site Management (Project Manager — 52 weeks @ €1,600/wk)", "wk", 52, 1600),
            ("1.2", "Site Manager — 52 weeks @ €1,400/wk", "wk", 52, 1400),
            ("1.3", "Welfare Facilities (cabins, sanitation)", "wk", 52, 300),
            ("1.4", "Performance Bond (10% of contract value @ 1.5%)", "item", 1, 0),
        ],
    },
    "uk-gc": {
        "currency_symbol": "£",
        "number_format": '£#,##0.00',
        "prelims_samples": [
            # UK CIJC-derived management rates (mid-2026 baseline; user should refresh)
            ("1.1", "Site Management (Contract Manager — 52 weeks @ £1,850/wk)", "wk", 52, 1850),
            ("1.2", "Site Manager — 52 weeks @ £1,550/wk", "wk", 52, 1550),
            ("1.3", "Welfare Facilities (cabins, sanitation)", "wk", 52, 320),
            ("1.4", "Performance Bond (10% of contract value @ 1.0%)", "item", 1, 0),
        ],
    },
}

# Populated by create_estimate_workbook() before any cells are written.
_ACTIVE_NUMBER_FORMAT = '€#,##0.00'
_ACTIVE_CURRENCY_SYMBOL = "€"

def header_style(cell, bg=AILTIR_NAVY):
    """Primary header row — navy background, white text, Space Grotesk."""
    cell.font = Font(bold=True, color=WHITE, size=11, name="Space Grotesk")
    cell.fill = PatternFill("solid", fgColor=bg)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = THIN_BORDER

def section_style(cell):
    """Section divider row — navy-700 background."""
    cell.font = Font(bold=True, color=WHITE, size=10, name="Space Grotesk")
    cell.fill = PatternFill("solid", fgColor=AILTIR_NAVY700)
    cell.border = THIN_BORDER

def total_style(cell):
    """Total/summary row — purple background."""
    cell.font = Font(bold=True, color=WHITE, size=10, name="Space Grotesk")
    cell.fill = PatternFill("solid", fgColor=AILTIR_PURPLE)
    cell.border = THIN_BORDER
    cell.number_format = _ACTIVE_NUMBER_FORMAT

def data_style(cell, bold=False, currency=False, alt_row=False):
    """Standard data cell."""
    cell.font = Font(bold=bold, size=10, name="Inter", color=AILTIR_NAVY)
    cell.alignment = Alignment(horizontal="right" if currency else "left", vertical="center")
    cell.border = THIN_BORDER
    if alt_row:
        cell.fill = PatternFill("solid", fgColor=AILTIR_LIGHT)
    if currency:
        cell.number_format = _ACTIVE_NUMBER_FORMAT

def amber_style(cell):
    """Amber highlight for key cost figures."""
    cell.font = Font(bold=True, size=10, name="Space Grotesk", color=AILTIR_NAVY)
    cell.fill = PatternFill("solid", fgColor=AILTIR_AMBER)
    cell.border = THIN_BORDER
    cell.number_format = _ACTIVE_NUMBER_FORMAT

def auto_width(ws):
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max_len + 4, 50)

def create_estimate_workbook(project_name, contract_value, output_path, profile_key="ireland-gc"):
    global _ACTIVE_NUMBER_FORMAT, _ACTIVE_CURRENCY_SYMBOL
    profile = PROFILE_DEFAULTS.get(profile_key)
    if profile is None:
        raise SystemExit(f"Unknown --profile-key: {profile_key!r}. Expected one of: {sorted(PROFILE_DEFAULTS)}")
    _ACTIVE_NUMBER_FORMAT = profile["number_format"]
    _ACTIVE_CURRENCY_SYMBOL = profile["currency_symbol"]

    wb = Workbook()

    # ── Sheet 1: Summary ──
    ws1 = wb.active
    ws1.title = "Summary"
    ws1.column_dimensions['A'].width = 35
    ws1.column_dimensions['B'].width = 20

    # Title block — navy header
    ws1.merge_cells("A1:B1")
    ws1['A1'] = "AILTIR — TENDER ESTIMATE SUMMARY"
    ws1['A1'].font = Font(bold=True, size=14, color=WHITE, name="Space Grotesk")
    ws1['A1'].fill = PatternFill("solid", fgColor=AILTIR_NAVY)
    ws1['A1'].alignment = Alignment(horizontal="center", vertical="center")

    ws1['A2'] = f"Project: {project_name}"
    ws1['A2'].font = Font(size=11, name="Inter", color=AILTIR_NAVY)
    ws1['A3'] = f"Date: {date.today().strftime('%d/%m/%Y')}"
    ws1['A3'].font = Font(size=11, name="Inter", color=AILTIR_NAVY)
    ws1['A4'] = "Status: DRAFT — Not for Submission"
    # Amber highlight for draft status — NOT red
    ws1['A4'].font = Font(bold=True, name="Space Grotesk", color=AILTIR_NAVY)
    ws1['A4'].fill = PatternFill("solid", fgColor=AILTIR_AMBER)

    ws1.append([])
    headers = ["Section", f"Amount ({_ACTIVE_CURRENCY_SYMBOL})"]
    ws1.append(headers)
    for cell in ws1[ws1.max_row]:
        header_style(cell)

    sections = [
        ("1. Preliminaries", "=Summary!B8"),
        ("2. Substructure / Groundworks", ""),
        ("3. Superstructure", ""),
        ("4. Envelope / Watertight", ""),
        ("5. Internal Finishes", ""),
        ("6. Mechanical & Electrical", ""),
        ("7. External Works", ""),
        ("DIRECT COSTS SUBTOTAL", "=SUM(B7:B13)"),
        ("", ""),
        ("Contingency (%)", ""),
        ("Margin (%)", ""),
        ("TOTAL TENDER PRICE", "=B14+B16+B17"),
    ]

    for i, (label, formula) in enumerate(sections):
        row = [label, formula if formula else 0]
        ws1.append(row)
        if "TOTAL TENDER" in label:
            for cell in ws1[ws1.max_row]:
                amber_style(cell)
        elif "SUBTOTAL" in label:
            for cell in ws1[ws1.max_row]:
                total_style(cell)
        else:
            data_style(ws1.cell(ws1.max_row, 1))
            data_style(ws1.cell(ws1.max_row, 2), currency=True)

    # ── Sheet 2: Priced Schedule ──
    ws2 = wb.create_sheet("Priced Schedule")
    ws2.column_dimensions['A'].width = 8
    ws2.column_dimensions['B'].width = 40
    ws2.column_dimensions['C'].width = 8
    ws2.column_dimensions['D'].width = 12
    ws2.column_dimensions['E'].width = 12
    ws2.column_dimensions['F'].width = 14

    ws2.append([
        "Item No",
        "Description",
        "Unit",
        "Qty",
        f"Rate ({_ACTIVE_CURRENCY_SYMBOL})",
        f"Amount ({_ACTIVE_CURRENCY_SYMBOL})",
    ])
    for cell in ws2[1]:
        header_style(cell)

    # Sample structure — NRM2 elemental format. Rates and rows come from PROFILE_DEFAULTS
    # so the same workbook generator emits an Irish or UK sample without a code change.
    prelims_rows = []
    for row_offset, (item_no, desc, unit, qty, rate) in enumerate(profile["prelims_samples"]):
        # Rows begin at spreadsheet row 3 (after the "1 PRELIMINARIES" header row on row 2).
        excel_row = 3 + row_offset
        formula = f"=D{excel_row}*E{excel_row}" if rate else "[CALCULATE]"
        prelims_rows.append((item_no, desc, unit, qty, rate, formula))

    sample_items = [
        ("1", "PRELIMINARIES", "", "", "", ""),
        *prelims_rows,
        ("2", "SUBSTRUCTURE", "", "", "", ""),
        ("2.1", "[Add items from takeoff]", "", "", "", ""),
        ("3", "SUPERSTRUCTURE", "", "", "", ""),
        ("3.1", "[Add items from takeoff]", "", "", "", ""),
    ]
    for item in sample_items:
        ws2.append(list(item))
        row = ws2.max_row
        # Section header rows — navy-700 fill
        if item[1].isupper() and item[0].isdigit() and len(item[0]) == 1:
            for cell in ws2[row]:
                section_style(cell)
        else:
            for cell in ws2[row]:
                cell.border = THIN_BORDER
        if item[5] and item[5] != "" and item[5] != "[CALCULATE]":
            ws2.cell(row, 6).number_format = _ACTIVE_NUMBER_FORMAT
            ws2.cell(row, 5).number_format = _ACTIVE_NUMBER_FORMAT

    # ── Sheet 3: Detailed Workings ──
    ws3 = wb.create_sheet("Workings")
    ws3.append([
        "Item No",
        "Description",
        "Pricing Basis",
        f"Labour {_ACTIVE_CURRENCY_SYMBOL}",
        f"Materials {_ACTIVE_CURRENCY_SYMBOL}",
        f"Plant {_ACTIVE_CURRENCY_SYMBOL}",
        f"Subbie {_ACTIVE_CURRENCY_SYMBOL}",
        f"Total {_ACTIVE_CURRENCY_SYMBOL}",
        "Source",
        "Assumptions",
    ])
    for cell in ws3[1]:
        header_style(cell)
    ws3.column_dimensions['B'].width = 35
    ws3.column_dimensions['C'].width = 25
    ws3.column_dimensions['I'].width = 20
    ws3.column_dimensions['J'].width = 30

    # ── Sheet 4: Subcontractor Register ──
    ws4 = wb.create_sheet("Subcontractor Register")
    ws4.append(["Subcontractor", "Trade", "Quote Ref", f"Amount ({_ACTIVE_CURRENCY_SYMBOL})", "Scope", "Exclusions", "Valid Until"])
    for cell in ws4[1]:
        header_style(cell)
    ws4.column_dimensions['A'].width = 25
    ws4.column_dimensions['E'].width = 35
    ws4.column_dimensions['F'].width = 35

    # ── Sheet 5: Assumptions Register ──
    ws5 = wb.create_sheet("Assumptions Register")
    ws5.append(["ASM ID", "Assumption", "Step", "Impact", "Status"])
    for cell in ws5[1]:
        header_style(cell)
    ws5.column_dimensions['B'].width = 50
    ws5.column_dimensions['D'].width = 25

    # ── Sheet 6: Rate Library Used ──
    ws6 = wb.create_sheet("Rates Used")
    ws6.append(["Rate Code", "Description", "Unit", f"Rate ({_ACTIVE_CURRENCY_SYMBOL})", "Applied To"])
    for cell in ws6[1]:
        header_style(cell)
    ws6.column_dimensions['B'].width = 35
    ws6.column_dimensions['E'].width = 30

    # Seed with current Irish SEO rates (Aug 2025 — Sectoral Employment Order)
    rates = [
        ("LAB-CRAFT", "Craftsperson (all-in incl. PRSI/pension)", "hr", 32.20, "Labour items"),
        ("LAB-CAT-A", "Category A Worker (all-in)", "hr", 31.25, "Labour items"),
        ("LAB-CAT-B", "Category B Worker (all-in)", "hr", 29.00, "Labour items"),
        ("MGMT-PM", "Project Manager", "wk", 1600, "Prelims"),
        ("MGMT-SM", "Site Manager", "wk", 1400, "Prelims"),
        ("MGMT-ENG", "Site Engineer", "wk", 1200, "Prelims"),
        ("MGMT-QS", "Quantity Surveyor (part-time)", "wk", 800, "Prelims"),
        ("SITE-CABIN", "Welfare Cabin (32ft)", "wk", 175, "Prelims"),
        ("SITE-HOARDING", "Hoarding (standard)", "m", 45, "Prelims"),
    ]
    for i, rate in enumerate(rates):
        ws6.append(list(rate))
        row = ws6.max_row
        if i % 2 == 1:
            for cell in ws6[row]:
                cell.fill = PatternFill("solid", fgColor=AILTIR_LIGHT)
        ws6.cell(row, 4).number_format = _ACTIVE_NUMBER_FORMAT

    auto_width(ws6)

    wb.save(output_path)
    print(f"✓ Estimate workbook created: {output_path}")

if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Create Ailtir Estimate Workbook")
    parser.add_argument("--project", default="New Project", help="Project name")
    parser.add_argument("--value", type=float, default=0, help="Estimated contract value")
    parser.add_argument("--output", default="estimate.xlsx", help="Output file path")
    parser.add_argument(
        "--profile-key",
        default="ireland-gc",
        choices=sorted(PROFILE_DEFAULTS.keys()),
        help="Ailtir profile controlling currency and sample prelims rates.",
    )
    args = parser.parse_args()

    create_estimate_workbook(args.project, args.value, args.output, profile_key=args.profile_key)
