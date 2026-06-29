#!/usr/bin/env python3
"""
Apply standard estimating styles to Excel workbooks.
"""

import sys
from pathlib import Path

def install_deps():
    try:
        from openpyxl import load_workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        import subprocess
        subprocess.run([sys.executable, "-m", "pip", "install", "openpyxl", "--break-system-packages"], check=True)

install_deps()

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, NamedStyle

# Ailtir Brand Colours
HEADER_FILL = PatternFill(start_color="0A1128", end_color="0A1128", fill_type="solid") # Navy 900
HEADER_FONT = Font(name="Space Grotesk", color="FFFFFF", bold=True)
SUBTOTAL_FILL = PatternFill(start_color="6D28D9", end_color="6D28D9", fill_type="solid") # Purple 600
SUBTOTAL_FONT = Font(name="Inter", color="FFFFFF", bold=True)
SECTION_FILL = PatternFill(start_color="1A2550", end_color="1A2550", fill_type="solid") # Navy 700
SECTION_FONT = Font(name="Space Grotesk", color="FFFFFF", bold=True)
BODY_FONT = Font(name="Inter", color="0A1128")
ALT_ROW_FILL = PatternFill(start_color="F5F7FA", end_color="F5F7FA", fill_type="solid")
CURRENCY_FORMAT = '€#,##0.00;[Red]-€#,##0.00'  # Red negatives = standard accounting convention, not brand colour
THIN_BORDER = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)


def style_header_row(ws, row: int = 1):
    """Apply header styling to a row."""
    for cell in ws[row]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)


def style_subtotal_row(ws, row: int):
    """Apply subtotal styling to a row."""
    for cell in ws[row]:
        cell.fill = SUBTOTAL_FILL
        cell.font = SUBTOTAL_FONT


def format_currency_column(ws, col: str):
    """Apply currency format to a column."""
    for cell in ws[col]:
        if cell.row > 1:  # Skip header
            cell.number_format = CURRENCY_FORMAT


def auto_column_width(ws, min_width: int = 10, max_width: int = 50):
    """Auto-fit column widths based on content."""
    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                cell_len = len(str(cell.value or ""))
                max_length = max(max_length, cell_len)
            except:
                pass
        adjusted = min(max(max_length + 2, min_width), max_width)
        ws.column_dimensions[col_letter].width = adjusted


def apply_borders_and_fonts(ws, start_row: int = 1, end_row: int = None):
    """Apply thin borders and body font to all cells, with alternating row colours."""
    if end_row is None:
        end_row = ws.max_row
    for i, row in enumerate(ws.iter_rows(min_row=start_row, max_row=end_row)):
        is_even_row = (i % 2 != 0) # 0-indexed, so this makes alternating rows
        for cell in row:
            cell.border = THIN_BORDER
            if cell.row > 1 and cell.fill.start_color.index == "00000000": # Don't overwrite headers
                cell.font = BODY_FONT
                if is_even_row:
                    cell.fill = ALT_ROW_FILL


def style_workbook(filepath: str, currency_cols: list = None, subtotal_rows: list = None):
    """Apply full styling to a workbook."""
    wb = load_workbook(filepath)
    
    for ws in wb.worksheets:
        style_header_row(ws)
        auto_column_width(ws)
        apply_borders_and_fonts(ws)
        
        if currency_cols:
            for col in currency_cols:
                format_currency_column(ws, col)
        
        if subtotal_rows:
            for row in subtotal_rows:
                style_subtotal_row(ws, row)
    
    wb.save(filepath)
    print(f"✓ Styled: {filepath}")


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Usage: python style_excel.py <file.xlsx> [--currency E,F] [--subtotals 10,20,30]")
        sys.exit(1)
    
    filepath = sys.argv[1]
    currency_cols = None
    subtotal_rows = None
    
    args = sys.argv[2:]
    for i, arg in enumerate(args):
        if arg == "--currency" and i + 1 < len(args):
            currency_cols = args[i + 1].split(",")
        elif arg == "--subtotals" and i + 1 < len(args):
            subtotal_rows = [int(r) for r in args[i + 1].split(",")]
    
    style_workbook(filepath, currency_cols, subtotal_rows)
