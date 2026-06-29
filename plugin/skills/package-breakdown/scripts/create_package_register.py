import argparse
import sys

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
except ImportError:
    print("openpyxl not installed. Please install it.")
    sys.exit(1)

AILTIR_NAVY = "0A1128" # Navy 900
AILTIR_WHITE = "FFFFFF"
HEADER_FONT = Font(bold=True, color=AILTIR_WHITE, size=11, name="Space Grotesk")
HEADER_FILL = PatternFill("solid", fgColor=AILTIR_NAVY)
THIN_BORDER = Border(
    left=Side(style="thin", color="D9D9D9"),
    right=Side(style="thin", color="D9D9D9"),
    top=Side(style="thin", color="D9D9D9"),
    bottom=Side(style="thin", color="D9D9D9"),
)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)

def style_header(ws, row, max_col):
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        cell.border = THIN_BORDER

def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--output", required=True)
    args = parser.parse_args()

    wb = Workbook()
    ws = wb.active
    ws.title = "Package Register"
    
    headers = ["Package ID", "Trade", "Estimated Value", "Target Issue Date", "Spec Sections", "Drawing Series", "Key Interfaces"]
    for i, h in enumerate(headers, 1):
        ws.cell(row=1, column=i, value=h)
    style_header(ws, 1, len(headers))
    
    wb.save(args.output)
    print(f"Created {args.output}")

if __name__ == "__main__":
    main()
