#!/usr/bin/env python3
"""
Custom Output Format Formatter for Construction Takeoffs

Allows contractors to upload their own Excel template and have the
takeoff output mapped into that format. If no template is provided,
the default excel_output.py format is used.

Workflow:
1. analyze_template() — reads a template Excel file, returns its structure
2. map_takeoff_to_template() — maps takeoff fields to template columns
3. format_output() — copies template and fills in takeoff data

This is a TAKEOFF tool — it populates quantities, units, descriptions,
and references. Rate/pricing columns are left blank for downstream
estimation skills.
"""

import sys
import json
import re
from copy import copy
from pathlib import Path

try:
    import openpyxl
    from openpyxl.utils import get_column_letter
except ImportError:
    print("ERROR: openpyxl required. Install with: pip install openpyxl", file=sys.stderr)
    sys.exit(1)


# Column name patterns for auto-mapping takeoff fields to template columns
COLUMN_PATTERNS = {
    "description": [
        r"(?i)desc", r"(?i)item.*(?:from|description)", r"(?i)scope",
        r"(?i)work.*item", r"(?i)element", r"(?i)activity",
    ],
    "quantity": [
        r"(?i)(?:number|no).*(?:of)?.*unit", r"(?i)qty", r"(?i)quant",
        r"(?i)measure", r"(?i)amount",
    ],
    "unit": [
        r"(?i)unit.*(?:of)?.*measure", r"(?i)\buom\b", r"(?i)\bunit\b",
    ],
    "trade": [
        r"(?i)trade", r"(?i)discipline", r"(?i)category", r"(?i)section",
    ],
    "drawing_ref": [
        r"(?i)draw", r"(?i)dwg", r"(?i)sheet.*ref", r"(?i)reference",
    ],
    "confidence": [
        r"(?i)confid", r"(?i)certainty", r"(?i)status",
    ],
    "tag": [
        r"(?i)\btag\b", r"(?i)\bcode\b", r"(?i)\bid\b", r"(?i)item.*no",
    ],
    "note": [
        r"(?i)note", r"(?i)comment", r"(?i)remark", r"(?i)assumption",
    ],
}

# Columns we should NOT populate (pricing — out of scope for takeoff)
PRICING_COLUMN_PATTERNS = [
    r"(?i)\brate\b", r"(?i)\bprice\b", r"(?i)\bcost\b",
    r"(?i)\btotal\b", r"(?i)\bamount\b", r"(?i)man.*hour",
    r"(?i)labour", r"(?i)labor", r"(?i)\b\$\b",
]


def analyze_template(template_path):
    """Read a template Excel file and return its structure.

    Args:
        template_path: path to the .xlsx template file

    Returns:
        dict with template structure:
        {
            "path": str,
            "sheets": [
                {
                    "name": str,
                    "header_row": int,
                    "data_start_row": int,
                    "columns": [
                        {"index": int, "letter": str, "name": str, "is_pricing": bool}
                    ],
                    "has_formulas": bool,
                    "sample_rows": int,
                    "merged_regions": list,
                }
            ],
            "primary_sheet": str (name of the main data sheet),
        }
    """
    wb = openpyxl.load_workbook(template_path, data_only=False)
    result = {
        "path": str(template_path),
        "sheets": [],
        "primary_sheet": None,
    }

    best_sheet = None
    best_col_count = 0

    for ws_name in wb.sheetnames:
        ws = wb[ws_name]
        sheet_info = {
            "name": ws_name,
            "header_row": None,
            "data_start_row": None,
            "columns": [],
            "has_formulas": False,
            "sample_rows": ws.max_row or 0,
            "merged_regions": [str(m) for m in ws.merged_cells.ranges],
        }

        # Find header row — the first row with 3+ non-empty cells
        for row_num in range(1, min(ws.max_row or 1, 20) + 1):
            non_empty = 0
            for col in range(1, (ws.max_column or 1) + 1):
                cell = ws.cell(row=row_num, column=col)
                if cell.value and str(cell.value).strip():
                    non_empty += 1
            if non_empty >= 3:
                sheet_info["header_row"] = row_num
                sheet_info["data_start_row"] = row_num + 1
                break

        # Extract column definitions from header row
        if sheet_info["header_row"]:
            hr = sheet_info["header_row"]
            for col in range(1, (ws.max_column or 1) + 1):
                cell = ws.cell(row=hr, column=col)
                col_name = str(cell.value).strip() if cell.value else ""
                if not col_name:
                    continue

                is_pricing = any(re.search(p, col_name) for p in PRICING_COLUMN_PATTERNS)

                sheet_info["columns"].append({
                    "index": col,
                    "letter": get_column_letter(col),
                    "name": col_name,
                    "is_pricing": is_pricing,
                })

        # Check for formulas
        for row_num in range(1, min(ws.max_row or 1, 50) + 1):
            for col in range(1, (ws.max_column or 1) + 1):
                cell = ws.cell(row=row_num, column=col)
                if isinstance(cell.value, str) and cell.value.startswith("="):
                    sheet_info["has_formulas"] = True
                    break
            if sheet_info["has_formulas"]:
                break

        result["sheets"].append(sheet_info)

        # Track which sheet has the most columns (likely the primary data sheet)
        if len(sheet_info["columns"]) > best_col_count:
            best_col_count = len(sheet_info["columns"])
            best_sheet = ws_name

    result["primary_sheet"] = best_sheet
    wb.close()

    return result


def auto_map_columns(template_structure, sheet_name=None):
    """Auto-map takeoff fields to template columns based on column name patterns.

    Args:
        template_structure: output from analyze_template()
        sheet_name: which sheet to map (defaults to primary_sheet)

    Returns:
        dict mapping takeoff field names to template column indices:
        {"description": 3, "quantity": 4, "unit": 5, ...}
    """
    if sheet_name is None:
        sheet_name = template_structure["primary_sheet"]

    sheet = None
    for s in template_structure["sheets"]:
        if s["name"] == sheet_name:
            sheet = s
            break

    if not sheet or not sheet["columns"]:
        return {}

    mapping = {}
    used_cols = set()

    for field, patterns in COLUMN_PATTERNS.items():
        for col_info in sheet["columns"]:
            if col_info["index"] in used_cols:
                continue
            col_name = col_info["name"]
            for pattern in patterns:
                if re.search(pattern, col_name):
                    mapping[field] = col_info["index"]
                    used_cols.add(col_info["index"])
                    break
            if field in mapping:
                break

    return mapping


def format_output(takeoff_data, template_path, output_path, mapping=None,
                  sheet_name=None, group_by_trade=True):
    """Copy the template and fill in takeoff data.

    Args:
        takeoff_data: list of quantity dicts (same format as excel_output.py expects)
        template_path: path to the user's .xlsx template
        output_path: where to save the populated output
        mapping: dict mapping field names to column indices (auto-detected if None)
        sheet_name: which sheet to populate (defaults to primary sheet)
        group_by_trade: whether to insert trade subheader rows

    Returns:
        str: path to the output file
    """
    # Analyze template if no mapping provided
    structure = analyze_template(template_path)
    if sheet_name is None:
        sheet_name = structure["primary_sheet"]

    sheet_info = None
    for s in structure["sheets"]:
        if s["name"] == sheet_name:
            sheet_info = s
            break

    if not sheet_info:
        print(f"ERROR: Sheet '{sheet_name}' not found in template", file=sys.stderr)
        return None

    if mapping is None:
        mapping = auto_map_columns(structure, sheet_name)

    if not mapping:
        print("ERROR: Could not auto-map any columns. Provide explicit mapping.", file=sys.stderr)
        return None

    # Load the template workbook (preserving formatting)
    wb = openpyxl.load_workbook(template_path)
    ws = wb[sheet_name]

    # Find where to start writing data
    data_start = sheet_info["data_start_row"] or 2

    # Clear existing data rows (but preserve formatting from first data row)
    first_data_row_styles = {}
    for col in range(1, (ws.max_column or 1) + 1):
        cell = ws.cell(row=data_start, column=col)
        first_data_row_styles[col] = {
            "font": copy(cell.font) if cell.font else None,
            "fill": copy(cell.fill) if cell.fill else None,
            "border": copy(cell.border) if cell.border else None,
            "alignment": copy(cell.alignment) if cell.alignment else None,
            "number_format": cell.number_format,
        }

    # Group quantities by trade if requested
    if group_by_trade:
        trades = []
        trade_items = {}
        for q in takeoff_data:
            trade = q.get("trade", "unknown")
            if trade not in trade_items:
                trades.append(trade)
                trade_items[trade] = []
            trade_items[trade].append(q)
        ordered_items = []
        for trade in trades:
            ordered_items.append({"_trade_header": trade})
            ordered_items.extend(trade_items[trade])
    else:
        ordered_items = takeoff_data

    # Write data
    row = data_start
    for item in ordered_items:
        # Trade subheader
        if "_trade_header" in item:
            if group_by_trade:
                # Merge across all mapped columns for the subheader
                max_col = max(mapping.values()) if mapping else 1
                ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=max_col)
                cell = ws.cell(row=row, column=1, value=item["_trade_header"].upper())
                cell.font = openpyxl.styles.Font(bold=True, name="Arial", size=10, color="1F4E79")
                cell.fill = openpyxl.styles.PatternFill("solid", fgColor="D6E4F0")
                row += 1
            continue

        # Write each mapped field
        for field, col_idx in mapping.items():
            value = None
            if field == "description":
                value = item.get("description", "")
            elif field == "quantity":
                value = item.get("quantity")
            elif field == "unit":
                value = item.get("unit", "")
            elif field == "trade":
                value = item.get("trade", "")
            elif field == "drawing_ref":
                value = item.get("drawing_ref", "")
            elif field == "confidence":
                value = item.get("confidence", "")
            elif field == "tag":
                value = item.get("tag", "")
            elif field == "note":
                value = item.get("note", "")

            if value is not None:
                cell = ws.cell(row=row, column=col_idx, value=value)
                # Apply formatting from the first data row
                style = first_data_row_styles.get(col_idx, {})
                if style.get("font"):
                    cell.font = copy(style["font"])
                if style.get("border"):
                    cell.border = copy(style["border"])
                if style.get("alignment"):
                    cell.alignment = copy(style["alignment"])
                if style.get("number_format"):
                    cell.number_format = style["number_format"]

        row += 1

    # Save
    output_path = str(output_path)
    wb.save(output_path)
    wb.close()
    print(f"Takeoff formatted to template: {output_path}", file=sys.stderr)

    return output_path


def describe_mapping(template_structure, mapping, sheet_name=None):
    """Generate a human-readable description of the column mapping.

    Used by Claude to show the user what will be mapped where before generating output.

    Returns:
        str: formatted mapping description
    """
    if sheet_name is None:
        sheet_name = template_structure["primary_sheet"]

    sheet = None
    for s in template_structure["sheets"]:
        if s["name"] == sheet_name:
            sheet = s
            break

    if not sheet:
        return "No sheet found."

    col_lookup = {c["index"]: c["name"] for c in sheet["columns"]}
    pricing_cols = [c["name"] for c in sheet["columns"] if c["is_pricing"]]

    lines = [f"Template: {Path(template_structure['path']).name}"]
    lines.append(f"Sheet: {sheet_name}")
    lines.append(f"Header row: {sheet['header_row']}")
    lines.append(f"Data starts: row {sheet['data_start_row']}")
    lines.append("")
    lines.append("FIELD MAPPING:")

    for field, col_idx in sorted(mapping.items(), key=lambda x: x[1]):
        col_name = col_lookup.get(col_idx, f"Col {col_idx}")
        lines.append(f"  {field:<15} → Column {get_column_letter(col_idx)} ({col_name})")

    if pricing_cols:
        lines.append("")
        lines.append("PRICING COLUMNS (left blank — out of scope for takeoff):")
        for col_name in pricing_cols:
            lines.append(f"  {col_name}")

    unmapped_cols = [c["name"] for c in sheet["columns"]
                     if c["index"] not in mapping.values() and not c["is_pricing"]]
    if unmapped_cols:
        lines.append("")
        lines.append("UNMAPPED COLUMNS (no takeoff field match):")
        for col_name in unmapped_cols:
            lines.append(f"  {col_name}")

    return "\n".join(lines)


# ── CLI ──

def main():
    """CLI for testing template analysis and formatting."""
    import argparse

    parser = argparse.ArgumentParser(description="Analyze or format to a custom Excel template")
    sub = parser.add_subparsers(dest="command")

    # analyze command
    p_analyze = sub.add_parser("analyze", help="Analyze a template structure")
    p_analyze.add_argument("template", help="Path to template .xlsx file")

    # format command
    p_format = sub.add_parser("format", help="Format takeoff data to a template")
    p_format.add_argument("template", help="Path to template .xlsx file")
    p_format.add_argument("data", help="Path to takeoff data JSON file")
    p_format.add_argument("-o", "--output", required=True, help="Output path")
    p_format.add_argument("--sheet", help="Sheet name to populate")

    args = parser.parse_args()

    if args.command == "analyze":
        structure = analyze_template(args.template)
        mapping = auto_map_columns(structure)
        print(describe_mapping(structure, mapping))
        print("\n--- RAW STRUCTURE ---")
        print(json.dumps(structure, indent=2))

    elif args.command == "format":
        with open(args.data) as f:
            data = json.load(f)
        format_output(data, args.template, args.output, sheet_name=args.sheet)

    else:
        parser.print_help()


if __name__ == "__main__":
    main()
