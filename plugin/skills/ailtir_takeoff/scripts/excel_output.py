"""Render a validated NRM2 Bill-of-Quantities JSON as an Ailtir-branded
Excel workbook (``Data`` + ``Flags`` sheets).

Brand styling is imported from the sibling skill
``ailtir_estimating-workflow/scripts/style_excel.py``. If that module is
not importable (single-file smoke run), a minimal in-file fallback with
the Ailtir palette is used so the workbook is still recognisably branded.

Sources referenced in comments (paraphrased):

* openpyxl docs — ``Workbook``, ``Worksheet.cell``, ``merge_cells``,
  string-formula HYPERLINK usage.
* RICS NRM2 — see ``research/nrm2-measurement.md`` §1 (element
  structure), §12 (description conventions), §14 (rounding).
* BS EN ISO 19650 role codes — ``research/drawing-conventions.md``
  §"Role codes" for the single-letter originator discipline field.
"""

from __future__ import annotations

import argparse
import json
import re
import sys
from dataclasses import dataclass
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Callable, Iterable, Optional, Sequence

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import column_index_from_string, get_column_letter
    from openpyxl.worksheet.worksheet import Worksheet
except ImportError:  # pragma: no cover
    sys.stderr.write(
        "[excel_output] fatal: openpyxl not installed. Install with:\n"
        "    python -m pip install --user openpyxl\n"
    )
    sys.exit(5)


# ---- style_excel import (primary) or inline fallback (safety net) --------
_STYLE_MOD = (Path(__file__).resolve().parents[2]
              / "ailtir_estimating-workflow" / "scripts")
if _STYLE_MOD.is_dir():
    sys.path.insert(0, str(_STYLE_MOD))

# Ailtir brand palette — kept as module-level constants either way so the
# rest of this file can reference them without branching.
NAVY_900 = "0A1128"     # Ailtir Navy 900   — primary header fill
NAVY_700 = "1A2550"     # Ailtir Navy 700   — section-row fill
PURPLE_600 = "6D28D9"   # Ailtir Purple 600 — subtotal-row fill
AMBER_400 = "F59E0B"    # Ailtir Amber 400  — warning/flag
LIGHT_BG = "F5F7FA"     # Alt-row body fill

try:  # pragma: no cover — thin import branch
    from style_excel import (  # type: ignore[import-not-found]
        apply_body_styling,
        auto_column_width,
        format_currency_column,
        style_header_row,
        style_section_row,
        style_subtotal_row,
    )
    _HAS_STYLE_EXCEL = True
except Exception as _style_exc:  # noqa: BLE001
    sys.stderr.write(
        f"[excel_output] warning: could not import style_excel "
        f"({_style_exc!r}); using inline brand-fallback.\n"
    )
    _HAS_STYLE_EXCEL = False

    _BORDER = Border(*(Side(border_style="thin", color="D9D9D9"),) * 4)

    def _fill(rgb: str) -> PatternFill:
        return PatternFill(fill_type="solid", start_color=rgb, end_color=rgb)

    def _heading_font() -> Font:
        return Font(name="Space Grotesk", size=11, bold=True, color="FFFFFF")

    def _apply_row(ws: Worksheet, row: int, rgb: str,
                   align: Alignment) -> None:
        fill, font = _fill(rgb), _heading_font()
        for col in range(1, (ws.max_column or 1) + 1):
            c = ws.cell(row=row, column=col)
            c.fill, c.font, c.alignment, c.border = fill, font, align, _BORDER

    def style_header_row(ws: Worksheet, row: int = 1) -> None:  # type: ignore[no-redef]
        _apply_row(ws, row, NAVY_900,
                   Alignment(horizontal="center", vertical="center",
                             wrap_text=True))

    def style_section_row(ws: Worksheet, row: int) -> None:  # type: ignore[no-redef]
        _apply_row(ws, row, NAVY_700,
                   Alignment(horizontal="left", vertical="center",
                             wrap_text=True))

    def style_subtotal_row(ws: Worksheet, row: int) -> None:  # type: ignore[no-redef]
        _apply_row(ws, row, PURPLE_600,
                   Alignment(horizontal="left", vertical="center",
                             wrap_text=True))

    def format_currency_column(ws: Worksheet, col_letter: str,  # type: ignore[no-redef]
                               profile_key: str) -> None:
        fmt = _CURRENCY_FORMATS.get(profile_key)
        if not fmt:
            return
        idx = column_index_from_string(col_letter)
        align = Alignment(horizontal="right", vertical="center")
        for row in range(2, (ws.max_row or 1) + 1):
            c = ws.cell(row=row, column=idx)
            c.number_format, c.alignment = fmt, align

    def auto_column_width(ws: Worksheet, min_width: int = 10,  # type: ignore[no-redef]
                          max_width: int = 60) -> None:
        for col_idx in range(1, (ws.max_column or 1) + 1):
            widest = min_width
            for row in range(1, (ws.max_row or 1) + 1):
                v = ws.cell(row=row, column=col_idx).value
                if v is None:
                    continue
                for line in str(v).splitlines() or [str(v)]:
                    widest = max(widest, len(line) + 2)
            ws.column_dimensions[get_column_letter(col_idx)].width = \
                min(widest, max_width)

    def apply_body_styling(ws: Worksheet, start_row: int = 2,  # type: ignore[no-redef]
                           end_row: Optional[int] = None,
                           alt_rows: bool = True) -> None:
        if end_row is None:
            end_row = ws.max_row or start_row
        if end_row < start_row:
            return
        font = Font(name="Inter", size=10, color="0A1128")
        align = Alignment(horizontal="left", vertical="top", wrap_text=True)
        alt_fill = _fill(LIGHT_BG)
        last = ws.max_column or 1
        for row in range(start_row, end_row + 1):
            shade = alt_rows and ((row - start_row) % 2 == 1)
            for col in range(1, last + 1):
                c = ws.cell(row=row, column=col)
                c.font, c.alignment, c.border = font, align, _BORDER
                if shade:
                    c.fill = alt_fill


# ---- Profile / currency --------------------------------------------------
_CURRENCY_FORMATS: dict[str, str] = {
    "ireland-gc": "€#,##0.00;[Red]-€#,##0.00",
    "uk-gc":      "£#,##0.00;[Red]-£#,##0.00",
}
_DEFAULT_PROFILE = "ireland-gc"


# ---- Column layout (letters are load-bearing — SUMIF & Total formulas) --
COL_ELEMENT_REF = "A"
COL_QUANTITY = "F"
COL_RATE = "K"
COL_TOTAL = "L"
_N_COLS = 12
DATA_HEADERS: tuple[str, ...] = (
    "Element Ref", "Element Name", "Item Type", "Description",
    "Unit", "Quantity", "Waste %", "Waste-adj Qty",
    "Confidence", "Source Ref", "Rate", "Total",
)
FLAGS_HEADERS: tuple[str, ...] = (
    "Flag Type", "Element Ref", "Description", "Quantity",
    "Reason", "Suggested Action", "Link",
)

# NRM2 rounding conventions — research file §14.
_QTY_FORMATS: dict[str, str] = {
    "m": "0.00", "m2": "0.00", "m²": "0.00", "m3": "0.00", "m³": "0.00",
    "nr": "0", "kg": "0", "tonne": "0.000", "item": "0", "sum": "0",
}
_VALID_UNITS = set(_QTY_FORMATS)
_ROLE_CODE_RE = re.compile(r"-([A-Z]{1,2})-")  # BS EN ISO 19650 role field


# ---- Small helpers -------------------------------------------------------
def _load_json(path: Path) -> dict[str, Any]:
    with path.open("r", encoding="utf-8") as fh:
        return json.load(fh)


def _resolve_profile(cli: Optional[str], boq: dict[str, Any]) -> str:
    """CLI → BoQ.project.profile_key → Context/profile.json → default."""
    if cli and cli != "auto":
        return cli
    proj_key = (boq.get("project") or {}).get("profile_key")
    if isinstance(proj_key, str) and proj_key:
        return proj_key
    ctx = Path("Context") / "profile.json"
    if ctx.is_file():
        try:
            key = json.loads(ctx.read_text(encoding="utf-8")).get("profile_key")
            if isinstance(key, str) and key:
                return key
        except Exception as exc:  # noqa: BLE001
            sys.stderr.write(
                f"[excel_output] warning: Context/profile.json unreadable "
                f"({exc!r}).\n"
            )
    sys.stderr.write(
        f"[excel_output] warning: no profile found; defaulting to "
        f"{_DEFAULT_PROFILE!r}.\n"
    )
    return _DEFAULT_PROFILE


def _normalise_unit(raw: Any) -> str:
    if raw is None:
        return ""
    s = str(raw).strip().lower()
    if not s:
        return ""
    if s in _VALID_UNITS:
        return s
    aliases = {"m^2": "m2", "m^3": "m3", "sq m": "m2", "cu m": "m3"}
    if s in aliases:
        return aliases[s]
    sys.stderr.write(
        f"[excel_output] warning: unrecognised NRM2 unit {raw!r}; "
        f"coerced to lowercase.\n"
    )
    return s


def _fmt_description(item: dict[str, Any]) -> str:
    """NRM2 principal-item + sub-classifications hierarchy (§12)."""
    parts: list[str] = []
    principal = str(item.get("principal_item") or "").strip()
    if principal:
        parts.append(principal)
    for sub in item.get("sub_classifications") or []:
        s = str(sub).strip()
        if s:
            parts.append(s)
    joined = "; ".join(parts)
    loc = str(item.get("location_qualifier") or "").strip()
    if loc:
        joined = f"{joined}; to {loc}" if joined else f"to {loc}"
    return joined


def _fmt_source_ref(item: dict[str, Any]) -> str:
    src = item.get("source_ref") or {}
    bits: list[str] = []
    drawing = str(src.get("drawing") or "").strip()
    if drawing:
        bits.append(drawing)
    page = src.get("page")
    if page not in (None, ""):
        bits.append(f"p.{page}")
    zone = str(src.get("zone") or "").strip()
    if zone:
        bits.append(zone)
    return " · ".join(bits)


def _fmt_confidence(raw: Any) -> str:
    if raw is None:
        return ""
    s = str(raw).strip().lower()
    return s.title() if s in {"high", "medium", "low"} else str(raw)


def _extract_discipline(drawing: str) -> Optional[str]:
    m = _ROLE_CODE_RE.search(drawing or "")
    return m.group(1) if m else None


def _element_disciplines(items: Iterable[dict[str, Any]]) -> list[str]:
    seen: list[str] = []
    for it in items:
        code = _extract_discipline(str((it.get("source_ref") or {})
                                       .get("drawing") or ""))
        if code and code not in seen:
            seen.append(code)
    return sorted(seen)


def _nrm1_sort_key(ref: str) -> tuple[int, ...]:
    """Numeric sort so `1.1, 1.2, 1.10, 2.1, 10.1` orders correctly."""
    out: list[int] = []
    for chunk in (ref or "").split("."):
        try:
            out.append(int(chunk))
        except ValueError:
            out.append(10_000)
    return tuple(out) if out else (10_000,)


# ---- Grouping ------------------------------------------------------------
@dataclass
class _ElementBlock:
    ref: str
    name: str
    items: list[dict[str, Any]]


def _group_by_element(items: Sequence[dict[str, Any]]) -> list[_ElementBlock]:
    buckets: dict[str, _ElementBlock] = {}
    for item in items:
        ref = str(item.get("element_ref") or "").strip() or "?"
        name = str(item.get("element_name") or "").strip()
        block = buckets.get(ref)
        if block is None:
            buckets[ref] = _ElementBlock(ref=ref, name=name, items=[item])
        else:
            if not block.name and name:
                block.name = name
            block.items.append(item)
    return sorted(buckets.values(), key=lambda b: _nrm1_sort_key(b.ref))


def _order_items(block: _ElementBlock) -> list[dict[str, Any]]:
    """Primary rows first; each primary's secondaries slotted beneath;
    orphaned secondaries appended at the end."""
    primaries: list[dict[str, Any]] = []
    secondaries: dict[str, list[dict[str, Any]]] = {}
    orphans: list[dict[str, Any]] = []
    for it in block.items:
        if str(it.get("item_type") or "primary").lower() == "primary":
            primaries.append(it)
        else:
            parent = it.get("parent_tag")
            (secondaries.setdefault(str(parent), []) if parent
             else orphans).append(it)
    ordered: list[dict[str, Any]] = []
    for prim in primaries:
        ordered.append(prim)
        ordered.extend(secondaries.pop(str(prim.get("tag") or ""), []))
    for leftover in secondaries.values():
        ordered.extend(leftover)
    ordered.extend(orphans)
    return ordered


# ---- Data sheet ----------------------------------------------------------
def _write_row(ws: Worksheet, row: int, item: dict[str, Any]) -> None:
    """Populate the 12-column body row. Rate (K) is blank; Total (L) is
    ``=IF(K<row>="","",F<row>*K<row>)`` — appears blank until priced."""
    unit = _normalise_unit(item.get("unit"))
    qty = item.get("quantity")
    waste = item.get("waste_pct")
    waste_adj = item.get("waste_adjusted_quantity")
    item_type = str(item.get("item_type") or "").strip().title() or "Primary"

    ws.cell(row=row, column=1,  value=str(item.get("element_ref") or ""))
    ws.cell(row=row, column=2,  value=str(item.get("element_name") or ""))
    ws.cell(row=row, column=3,  value=item_type)
    ws.cell(row=row, column=4,  value=_fmt_description(item))
    ws.cell(row=row, column=5,  value=unit)
    qty_cell = ws.cell(row=row, column=6,
                       value=qty if isinstance(qty, (int, float)) else None)
    waste_cell = ws.cell(row=row, column=7,
                         value=waste if isinstance(waste, (int, float)) else None)
    waste_adj_cell = ws.cell(row=row, column=8,
                             value=waste_adj if isinstance(waste_adj, (int, float)) else None)
    ws.cell(row=row, column=9,  value=_fmt_confidence(item.get("confidence")))
    ws.cell(row=row, column=10, value=_fmt_source_ref(item))
    ws.cell(row=row, column=11, value=None)  # Rate — blank
    ws.cell(
        row=row, column=12,
        value=f'=IF({COL_RATE}{row}="","",{COL_QUANTITY}{row}*{COL_RATE}{row})',
    )
    qty_fmt = _QTY_FORMATS.get(unit, "0.00")
    if qty_cell.value is not None:
        qty_cell.number_format = qty_fmt
    if waste_adj_cell.value is not None:
        waste_adj_cell.number_format = qty_fmt
    if waste_cell.value is not None:
        waste_cell.number_format = '0.0"%"'


def _render_data_sheet(ws: Worksheet, boq: dict[str, Any],
                       profile_key: str) -> dict[str, int]:
    """Populate Sheet 1. Returns ``{item_tag: row_index}`` for the Flags
    sheet's HYPERLINK targets."""
    project = boq.get("project") or {}
    project_name = str(project.get("name") or "Untitled project")
    src_pdfs = project.get("source_pdfs") or []
    page_count = project.get("page_count") or 0
    generated_at = str(project.get("generated_at") or "")
    gen_date = (generated_at.split("T", 1)[0] if generated_at else
                datetime.now(timezone.utc).strftime("%Y-%m-%d"))

    ws.title = "Data"
    ws.cell(row=1, column=1,
            value=f"AILTIR NRM2 TAKEOFF REGISTER — {project_name}")
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=_N_COLS)
    style_header_row(ws, row=1)

    meta = (f"Profile: {profile_key} | Generated: {gen_date} | "
            f"Source: {len(src_pdfs)} PDF(s), {page_count} pages")
    meta_cell = ws.cell(row=2, column=1, value=meta)
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=_N_COLS)
    meta_cell.font = Font(name="Inter", size=9, italic=True)
    meta_cell.alignment = Alignment(horizontal="left", vertical="center")

    for idx, header in enumerate(DATA_HEADERS, start=1):
        ws.cell(row=4, column=idx, value=header)
    style_header_row(ws, row=4)

    tag_to_row: dict[str, int] = {}
    element_total_rows: list[int] = []
    items = [i for i in (boq.get("items") or []) if isinstance(i, dict)]

    current = 5
    body_start = current
    for block in _group_by_element(items):
        disciplines = _element_disciplines(block.items)
        heading = f"Element {block.ref} — {block.name}"
        if len(disciplines) > 1:
            heading += f" (disciplines: {', '.join(disciplines)})"
        ws.cell(row=current, column=1, value=heading)
        ws.merge_cells(start_row=current, start_column=1,
                       end_row=current, end_column=_N_COLS)
        style_section_row(ws, row=current)
        current += 1

        for item in _order_items(block):
            try:
                _write_row(ws, current, item)
            except Exception as exc:  # noqa: BLE001 — one bad row must not abort
                sys.stderr.write(
                    f"[excel_output] warning: could not render item "
                    f"{item.get('tag') or item.get('element_ref')!r}: "
                    f"{exc!r}. Skipping.\n"
                )
                for c in range(1, _N_COLS + 1):
                    ws.cell(row=current, column=c, value=None)
                current += 1
                continue

            tag = item.get("tag")
            if isinstance(tag, str) and tag:
                tag_to_row[tag] = current
            # Primary rows get the Purple-600 subtotal (parent-line)
            # treatment; secondaries stay body-styled.
            if str(item.get("item_type") or "").lower() == "primary":
                style_subtotal_row(ws, row=current)
            current += 1

        # Element total row — SUMIF over column A with a wildcard so
        # deeper sub-refs (e.g. 1.1.1) still roll up into element 1.1.
        ws.cell(row=current, column=1, value=f"Element {block.ref} total")
        ws.merge_cells(start_row=current, start_column=1,
                       end_row=current, end_column=10)
        ws.cell(
            row=current, column=12,
            value=(f'=SUMIF({COL_ELEMENT_REF}:{COL_ELEMENT_REF},'
                   f'"{block.ref}*",{COL_TOTAL}:{COL_TOTAL})'),
        )
        style_subtotal_row(ws, row=current)
        element_total_rows.append(current)
        current += 1

    if element_total_rows:
        ws.cell(row=current, column=1, value="TAKEOFF SUBTOTAL")
        ws.merge_cells(start_row=current, start_column=1,
                       end_row=current, end_column=10)
        refs = ",".join(f"{COL_TOTAL}{r}" for r in element_total_rows)
        ws.cell(row=current, column=12, value=f"=SUM({refs})")
        style_subtotal_row(ws, row=current)
        current += 1

    body_end = current - 1
    if body_end >= body_start:
        apply_body_styling(ws, start_row=body_start, end_row=body_end,
                           alt_rows=True)
        # Re-assert section / subtotal fills on top of the alt-row shading.
        for r in element_total_rows:
            style_subtotal_row(ws, row=r)

    format_currency_column(ws, COL_RATE, profile_key)
    format_currency_column(ws, COL_TOTAL, profile_key)
    auto_column_width(ws, min_width=10, max_width=60)
    style_header_row(ws, row=4)
    return tag_to_row


# ---- Flags sheet ---------------------------------------------------------
def _render_flags_sheet(ws: Worksheet, boq: dict[str, Any],
                        tag_to_row: dict[str, int]) -> None:
    project_name = str((boq.get("project") or {}).get("name")
                       or "Untitled project")
    ws.title = "Flags"
    ws.cell(row=1, column=1, value=f"REVIEW FLAGS — {project_name}")
    ws.merge_cells(start_row=1, start_column=1, end_row=1,
                   end_column=len(FLAGS_HEADERS))
    style_header_row(ws, row=1)

    flags = [f for f in (boq.get("flags") or []) if isinstance(f, dict)]
    element_count = len({str(f.get("element_ref") or "") for f in flags
                         if f.get("element_ref")})
    meta_cell = ws.cell(
        row=2, column=1,
        value=f"{len(flags)} flags across {element_count} elements",
    )
    ws.merge_cells(start_row=2, start_column=1, end_row=2,
                   end_column=len(FLAGS_HEADERS))
    meta_cell.font = Font(name="Inter", size=9, italic=True)
    meta_cell.alignment = Alignment(horizontal="left", vertical="center")

    for idx, header in enumerate(FLAGS_HEADERS, start=1):
        ws.cell(row=4, column=idx, value=header)
    style_header_row(ws, row=4)

    if not flags:
        ws.cell(row=5, column=1, value="No flags — review clean.")
        ws.merge_cells(start_row=5, start_column=1, end_row=5,
                       end_column=len(FLAGS_HEADERS))
        auto_column_width(ws, min_width=10, max_width=60)
        return

    sorted_flags = sorted(
        flags,
        key=lambda f: (_nrm1_sort_key(str(f.get("element_ref") or "")),
                       str(f.get("type") or "")),
    )
    row = 5
    for flag in sorted_flags:
        ws.cell(row=row, column=1, value=str(flag.get("type") or ""))
        ws.cell(row=row, column=2, value=str(flag.get("element_ref") or ""))
        ws.cell(row=row, column=3, value=str(flag.get("description") or ""))
        qty = flag.get("quantity")
        ws.cell(row=row, column=4,
                value=qty if isinstance(qty, (int, float)) else None)
        ws.cell(row=row, column=5, value=str(flag.get("reason") or ""))
        ws.cell(row=row, column=6,
                value=str(flag.get("suggested_action") or ""))
        item_tag = flag.get("item_tag")
        if isinstance(item_tag, str) and item_tag in tag_to_row:
            # openpyxl string-formula HYPERLINK renders as a clickable
            # anchor pointing at Data!A<row>.
            target = tag_to_row[item_tag]
            ws.cell(row=row, column=7,
                    value=f'=HYPERLINK("#Data!A{target}","Go to row")')
        row += 1

    apply_body_styling(ws, start_row=5, end_row=row - 1, alt_rows=True)
    auto_column_width(ws, min_width=10, max_width=60)
    style_header_row(ws, row=4)


# ---- Template-mode branch ------------------------------------------------
def _run_template_mode(template_path: Path, boq: dict[str, Any],
                       out_path: Path, quiet: bool) -> int:
    if not template_path.is_file():
        sys.stderr.write(
            f"[excel_output] error: template not found: {template_path}\n"
        )
        return 4
    try:
        import template_formatter  # type: ignore[import-not-found]
    except Exception as exc:  # noqa: BLE001
        sys.stderr.write(
            f"[excel_output] error: template_formatter import failed "
            f"({exc!r}).\n"
        )
        return 4
    render: Optional[Callable[..., int]] = getattr(
        template_formatter, "render", None,
    ) or getattr(template_formatter, "format_output", None)
    if not callable(render):
        sys.stderr.write(
            "[excel_output] error: template_formatter exposes neither "
            "render() nor format_output().\n"
        )
        return 4
    try:
        rc = render(template_path=str(template_path), boq=boq,
                    out_path=str(out_path))
    except TypeError:
        rc = render(template_path=str(template_path), boq_json=boq,
                    output_path=str(out_path), auto_map=True)
    rc = int(rc) if isinstance(rc, int) else 0
    if rc == 0 and not quiet:
        print(f"[excel_output] template-mode workbook written: {out_path}")
    return rc


# ---- CLI -----------------------------------------------------------------
def _build_parser() -> argparse.ArgumentParser:
    p = argparse.ArgumentParser(
        prog="excel_output.py",
        description="Render a validated NRM2 BoQ JSON as an Ailtir-branded "
                    "Excel workbook (Data + Flags sheets).",
    )
    p.add_argument("--json", required=True,
                   help="Path to validated BoQ JSON.")
    p.add_argument("--out", required=True,
                   help="Path for the generated .xlsx file.")
    p.add_argument("--profile", required=True,
                   choices=sorted(_CURRENCY_FORMATS.keys()) + ["auto"],
                   help="Ailtir profile key. 'auto' reads "
                        "Context/profile.json.")
    p.add_argument("--template", default=None,
                   help="Optional .xlsx template — delegates to "
                        "template_formatter.")
    p.add_argument("--quiet", action="store_true",
                   help="Suppress non-warning stdout.")
    return p


def main(argv: Optional[Sequence[str]] = None) -> int:
    args = _build_parser().parse_args(argv)
    json_path, out_path = Path(args.json), Path(args.out)

    if not json_path.is_file():
        sys.stderr.write(
            f"[excel_output] error: JSON not found: {json_path}\n"
        )
        return 2
    try:
        boq = _load_json(json_path)
        if not isinstance(boq, dict):
            raise ValueError("top-level JSON is not an object")
    except Exception as exc:  # noqa: BLE001
        sys.stderr.write(
            f"[excel_output] error: BoQ JSON unreadable: {exc!r}\n"
        )
        return 2

    try:
        out_path.parent.mkdir(parents=True, exist_ok=True)
    except OSError as exc:
        sys.stderr.write(
            f"[excel_output] error: cannot create {out_path.parent}: "
            f"{exc!r}\n"
        )
        return 3

    profile_key = _resolve_profile(args.profile, boq)
    if profile_key not in _CURRENCY_FORMATS:
        sys.stderr.write(
            f"[excel_output] warning: unknown profile {profile_key!r}; "
            f"currency formatting will be skipped.\n"
        )

    if args.template:
        return _run_template_mode(Path(args.template), boq, out_path,
                                  quiet=args.quiet)

    items = [i for i in (boq.get("items") or []) if isinstance(i, dict)]
    if not items:
        sys.stderr.write(
            "[excel_output] error: BoQ contains zero items; refusing to "
            "emit an empty workbook.\n"
        )
        return 6

    wb = Workbook()
    tag_to_row = _render_data_sheet(wb.active, boq, profile_key)
    _render_flags_sheet(wb.create_sheet(title="Flags"), boq, tag_to_row)

    try:
        wb.save(str(out_path))
    except OSError as exc:
        sys.stderr.write(
            f"[excel_output] error: cannot write {out_path}: {exc!r}\n"
        )
        return 3

    if not args.quiet:
        style_note = ("shared style module" if _HAS_STYLE_EXCEL
                      else "inline brand fallback")
        print(
            f"[excel_output] wrote {out_path} "
            f"(profile={profile_key}, items={len(items)}, "
            f"styling={style_note})"
        )
    return 0


if __name__ == "__main__":
    sys.exit(main())
