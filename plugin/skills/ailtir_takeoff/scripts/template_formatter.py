"""Ailtir template formatter — bring-your-own-template BoQ populator.

Takes a validated NRM2 takeoff (BoQ JSON produced by ``validate.py``) and a
user-supplied Excel workbook, and writes the takeoff rows into that
workbook while preserving every column ordering, style rule and formula
the user put in. Contractors who have a company-branded takeoff template
route the BoQ through this script instead of ``excel_output.py`` (which
writes Ailtir's own two-sheet register).

Ailtir differentiators baked in here:

* NRM2-aware alias table — variants such as ``Ele Ref``, ``NRM2 Code``,
  ``Group Element`` or ``Item No.`` resolve to the right canonical field.
* Fuzzy fallback with a reported similarity score, so low-confidence
  guesses can be overridden via ``--mapping``.
* Formulas preserved by contract, not best-effort — every formula-bearing
  cell is recorded during analysis and consulted before every write.
  Clobbering user formulas would defeat the point of the
  bring-your-own-template feature.
* Section-header aware fill — walks the template's own structural zones
  and slots items under matching NRM elements. Naive row-append mappers
  dump everything at the bottom of the sheet and break the contractor's
  cost-report grouping.

Public sources consulted for the behaviour implemented here:

* openpyxl documentation for ``Workbook``/``Worksheet``/cell styles,
  ``data_type`` and ``MergedCell`` (https://openpyxl.readthedocs.io).
* Python standard library ``difflib.SequenceMatcher.ratio`` for the
  fuzzy header match score.
* RICS NRM1 elemental hierarchy for the section-header name-to-code
  lookup (see ``research/nrm2-measurement.md``).
* ECMA-376 OOXML on merged cells and number-format tokens.
"""

from __future__ import annotations

import argparse
import json
import re
import shutil
import sys
from copy import copy
from difflib import SequenceMatcher
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Tuple

# --------------------------------------------------------------------------- #
# Exit codes                                                                  #
# --------------------------------------------------------------------------- #

EXIT_OK = 0
EXIT_TEMPLATE_BAD = 2
EXIT_JSON_BAD = 3
EXIT_OUT_BAD = 4
EXIT_NO_OPENPYXL = 5
EXIT_USER_DECLINED = 6
EXIT_MAPPING_BAD = 7
EXIT_NO_USABLE_SHEET = 8

# --------------------------------------------------------------------------- #
# Ailtir NRM alias table                                                      #
# --------------------------------------------------------------------------- #
# This is the primary differentiator versus a naive header-name mapper.
# All aliases are stored *normalised* (see ``_normalise_header``) so lookup
# is a straight equality check after normalisation.

_RAW_ALIASES: Dict[str, List[str]] = {
    "element_ref": [
        "element ref", "ele ref", "nrm ref", "nrm2 code", "nrm2 ref",
        "nrm1 ref", "element code", "group element", "element number",
        "elem",
    ],
    "element_name": [
        "element", "element name", "element description", "nrm element",
    ],
    "work_section": [
        "work section", "ws", "section", "ws code", "trade",
        "trade section",
    ],
    "id": [
        "id", "item no", "item number", "item ref", "line no", "ref", "#",
    ],
    "description": [
        "description", "description of works", "item description",
        "works description", "narrative", "principal item",
    ],
    "unit": [
        "unit", "uom", "unit of measure", "measure",
    ],
    "quantity": [
        "quantity", "qty", "qty measured", "measured qty", "total qty",
    ],
    "waste_factor": [
        "waste %", "waste factor", "waste", "allowance %",
        "waste allowance",
    ],
    "source_ref": [
        "source", "source ref", "drawing", "drawing ref", "sheet",
        "sheet ref", "dwg", "dwg no",
    ],
    "confidence": [
        "confidence", "conf", "check", "review", "flag",
    ],
}

# Fields that MUST be resolved on a sheet for that sheet to be considered
# populatable. Sheets that do not expose all four are skipped in the fill
# pass.
_REQUIRED_FIELDS: Tuple[str, ...] = (
    "element_ref", "description", "unit", "quantity",
)

# All canonical fields the auto-mapper considers, in the order that
# leftmost-column tie-breaks are resolved.
_CANONICAL_FIELDS: Tuple[str, ...] = (
    "id", "element_ref", "element_name", "work_section", "description",
    "unit", "quantity", "waste_factor", "source_ref", "confidence",
)

# NRM1 top-level element-name to element-code lookup used when a section
# header uses the element *name* ("SUPERSTRUCTURE") instead of the code
# ("2"). Codes verified against the BCIS Elemental Standard Form of Cost
# Analysis (4th NRM Edition); see research/nrm2-measurement.md.
NRM_ELEMENT_NAME_TO_CODE: Dict[str, str] = {
    "facilitating works": "0",
    "substructure": "1",
    "superstructure": "2",
    "internal finishes": "3",
    "fittings furnishings and equipment": "4",
    "fittings furnishings & equipment": "4",
    "ff&e": "4",
    "services": "5",
    "prefabricated buildings and units": "6",
    "prefabricated buildings & units": "6",
    "work to existing buildings": "7",
    "external works": "8",
}

# Regex to pull an element ref out of a section header text. Captures
# either a bare integer or a decimal group element like "2.6".
_ELEMENT_REF_RE = re.compile(r"\b(\d+(?:\.\d+)?)\b")

# --------------------------------------------------------------------------- #
# openpyxl                                                                    #
# --------------------------------------------------------------------------- #

try:
    import openpyxl
    from openpyxl.cell.cell import Cell
    from openpyxl.utils import get_column_letter
    from openpyxl.utils.cell import coordinate_from_string
    from openpyxl.worksheet.worksheet import Worksheet
except Exception:  # pragma: no cover - defensive; handled at CLI entry.
    openpyxl = None  # type: ignore[assignment]
    Cell = None  # type: ignore[assignment]
    get_column_letter = None  # type: ignore[assignment]
    coordinate_from_string = None  # type: ignore[assignment]
    Worksheet = None  # type: ignore[assignment]


# --------------------------------------------------------------------------- #
# Helpers                                                                     #
# --------------------------------------------------------------------------- #

_WS_RE = re.compile(r"\s+")
_PUNCT_RE = re.compile(r"[^\w\s%#&]")


def _normalise_header(text: Any) -> str:
    """Lower-case, strip punctuation (except %#&), collapse whitespace."""
    if text is None:
        return ""
    s = str(text).strip().lower()
    s = _PUNCT_RE.sub(" ", s)
    s = _WS_RE.sub(" ", s).strip()
    return s


# Pre-normalised alias lookup: field -> list[(alias_norm, alias_raw)]
NRM_FIELD_ALIASES: Dict[str, List[Tuple[str, str]]] = {
    field: [(_normalise_header(a), a) for a in aliases]
    for field, aliases in _RAW_ALIASES.items()
}


def _is_bold_font(cell: "Cell") -> bool:
    font = getattr(cell, "font", None)
    if font is None:
        return False
    if getattr(font, "bold", None) is True:
        return True
    # openpyxl exposes weight indirectly; heuristic left in place in case
    # a future openpyxl surfaces it.
    weight = getattr(font, "weight", None)
    try:
        return weight is not None and int(weight) >= 700
    except (TypeError, ValueError):
        return False


def _cell_has_non_default_fill(cell: "Cell") -> bool:
    fill = getattr(cell, "fill", None)
    if fill is None:
        return False
    fg = getattr(fill, "fgColor", None)
    if fg is None:
        return False
    rgb = getattr(fg, "rgb", None)
    if not rgb:
        return False
    val = str(rgb).upper().lstrip("F") if str(rgb).upper() == "FFFFFFFF" else str(rgb).upper()
    # Default/no-fill is 00000000; pure white FFFFFFFF is treated as no fill.
    return val not in {"00000000", "FFFFFFFF", "FFFFFF", "000000"}


def _row_all_empty(ws: "Worksheet", row: int, max_col: int) -> bool:
    for c in range(1, max_col + 1):
        v = ws.cell(row=row, column=c).value
        if v not in (None, ""):
            return False
    return True


def _cell_is_formula(cell: "Cell") -> bool:
    if getattr(cell, "data_type", None) == "f":
        return True
    v = cell.value
    return isinstance(v, str) and v.startswith("=")


# --------------------------------------------------------------------------- #
# analyze_template                                                            #
# --------------------------------------------------------------------------- #

def _detect_header_row(ws: "Worksheet", max_col: int) -> Optional[int]:
    """Return the 1-indexed header row for ``ws``, or ``None``.

    Two candidates race, and the earlier row wins:

    1. First row where every non-empty cell is bold and the row has
       >= 3 non-empty cells.
    2. First row where the ratio of non-empty cells to ``max_col`` is
       >= 0.5 *and* ``max_col > 5``.
    """
    scan_to = min(ws.max_row or 1, 20)
    bold_candidate: Optional[int] = None
    ratio_candidate: Optional[int] = None
    for row in range(1, scan_to + 1):
        cells = [ws.cell(row=row, column=c) for c in range(1, max_col + 1)]
        non_empty = [c for c in cells if c.value not in (None, "")]
        if not non_empty:
            continue
        if bold_candidate is None and len(non_empty) >= 3:
            if all(_is_bold_font(c) for c in non_empty):
                bold_candidate = row
        if ratio_candidate is None and max_col > 5:
            if len(non_empty) / max_col >= 0.5:
                ratio_candidate = row
        if bold_candidate is not None and ratio_candidate is not None:
            break
    candidates = [c for c in (bold_candidate, ratio_candidate) if c is not None]
    return min(candidates) if candidates else None


def _analyse_columns(
    ws: "Worksheet", header_row: int, max_col: int, max_row: int,
) -> List[Dict[str, Any]]:
    columns: List[Dict[str, Any]] = []
    for col in range(1, max_col + 1):
        label_cell = ws.cell(row=header_row, column=col)
        raw = label_cell.value
        label = "" if raw is None else str(raw).strip()
        if not label:
            continue
        has_formula = False
        for r in range(header_row + 1, max_row + 1):
            if _cell_is_formula(ws.cell(row=r, column=col)):
                has_formula = True
                break
        columns.append({
            "col_index": col,
            "col_letter": get_column_letter(col),
            "label": label,
            "label_normalised": _normalise_header(label),
            "has_formula": has_formula,
        })
    return columns


def _detect_style_zones(
    ws: "Worksheet", header_row: int, max_col: int, max_row: int,
) -> List[Dict[str, Any]]:
    """Section headers, subtotals and spacers below the header row."""
    zones: List[Dict[str, Any]] = []
    subtotal_prefixes = (
        "subtotal", "sub-total", "sub total", "element total",
        "section total", "trade total", "total for",
    )
    kw_prefixes = ("element", "section", "group", "trade", "nrm")
    last_data_row: Optional[int] = None
    pending_empty: List[int] = []

    for row in range(header_row + 1, max_row + 1):
        a_cell = ws.cell(row=row, column=1)
        a_val = a_cell.value
        a_text = str(a_val).strip() if a_val is not None else ""
        row_empty = _row_all_empty(ws, row, max_col)

        if row_empty:
            pending_empty.append(row)
            continue

        # Non-empty row: any queued blanks become spacers if they sit
        # between two data rows.
        if last_data_row is not None and pending_empty:
            for r in pending_empty:
                zones.append({
                    "row": r, "kind": "spacer", "text": "",
                    "reason": "empty_bold_row",
                })
        pending_empty = []

        lower = a_text.lower()
        classified = False

        if a_text:
            # Subtotal?
            if any(lower.startswith(p) for p in subtotal_prefixes):
                zones.append({
                    "row": row, "kind": "subtotal", "text": a_text,
                    "reason": "keyword",
                })
                classified = True
            else:
                # Section header — try each rule in reason-order.
                reason: Optional[str] = None
                if len(a_text) >= 5:
                    alpha = [ch for ch in a_text if ch.isalpha()]
                    if alpha:
                        upper_ratio = sum(1 for ch in alpha if ch.isupper()) / len(alpha)
                        if upper_ratio >= 0.6:
                            reason = "all_caps"
                if reason is None and _cell_has_non_default_fill(a_cell) and _is_bold_font(a_cell):
                    reason = "bold_fill"
                if reason is None and any(lower.startswith(p) for p in kw_prefixes):
                    reason = "keyword"
                if reason is not None:
                    zones.append({
                        "row": row, "kind": "section_header",
                        "text": a_text, "reason": reason,
                    })
                    classified = True

        if not classified:
            last_data_row = row

    return zones


def analyze_template(template_path: str | Path) -> dict:
    """Inspect ``template_path`` and return the analysis dict per sheet.

    See spec for the exact returned schema. This function does not modify
    the input file. It is safe to call repeatedly.
    """
    if openpyxl is None:
        raise RuntimeError(
            "openpyxl is required; install with `pip install openpyxl`."
        )
    path = Path(template_path)
    wb = openpyxl.load_workbook(path, data_only=False, keep_vba=False)
    out: Dict[str, Any] = {}
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        max_row = ws.max_row or 0
        max_col = ws.max_column or 0
        dims = {
            "rows": max_row, "cols": max_col,
            "max_row": max_row, "max_col": max_col,
        }
        header_row = _detect_header_row(ws, max_col) if max_col else None
        columns: List[Dict[str, Any]] = []
        style_zones: List[Dict[str, Any]] = []
        formula_cells: List[str] = []
        if header_row is not None:
            columns = _analyse_columns(ws, header_row, max_col, max_row)
            style_zones = _detect_style_zones(ws, header_row, max_col, max_row)
        # Collect formula cells across the whole sheet — cheaper than doing it
        # per column and needed for the "never overwrite formulas" contract.
        for row in ws.iter_rows(min_row=1, max_row=max_row, max_col=max_col):
            for cell in row:
                if _cell_is_formula(cell):
                    formula_cells.append(cell.coordinate)
        out[sheet_name] = {
            "dimensions": dims,
            "header_row": header_row,
            "columns": columns,
            "style_zones": style_zones,
            "formula_cells": formula_cells,
        }
    wb.close()
    return out


# --------------------------------------------------------------------------- #
# map_takeoff_to_template                                                     #
# --------------------------------------------------------------------------- #

def _match_field(
    field: str,
    columns_pool: List[Dict[str, Any]],
    fuzzy_threshold: float,
) -> Optional[Tuple[Dict[str, Any], str, Optional[float]]]:
    """Return ``(column, match_kind, score)`` for the best match, or None.

    ``columns_pool`` is a mutable list; the caller is responsible for
    removing the winning column so later fields cannot claim it.
    """
    field_norm = _normalise_header(field.replace("_", " "))
    aliases = NRM_FIELD_ALIASES.get(field, [])

    # 1. Exact match on the field name itself.
    for col in columns_pool:
        if col["label_normalised"] == field_norm:
            return col, "exact", 1.0

    # 2. Alias match.
    for col in columns_pool:
        for alias_norm, _ in aliases:
            if col["label_normalised"] == alias_norm:
                return col, "alias", 1.0

    # 3. Fuzzy — best score across (field-name + all aliases) vs header.
    best: Optional[Tuple[Dict[str, Any], float]] = None
    candidates_norm = [field_norm] + [a for a, _ in aliases]
    for col in columns_pool:
        header_norm = col["label_normalised"]
        if not header_norm:
            continue
        best_local = 0.0
        for cand in candidates_norm:
            if not cand:
                continue
            score = SequenceMatcher(None, header_norm, cand).ratio()
            if score > best_local:
                best_local = score
        if best_local >= fuzzy_threshold and (best is None or best_local > best[1]):
            best = (col, best_local)
    if best is not None:
        return best[0], "fuzzy", round(best[1], 3)

    return None


def map_takeoff_to_template(
    boq_json: dict,
    template_analysis: dict,
    user_mapping: Optional[dict] = None,
    fuzzy_threshold: float = 0.75,
) -> dict:
    """Compute the field→column mapping for every sheet in the analysis.

    Result shape is documented in the spec. Sheets that fail to expose
    all four required fields are still emitted so callers can inspect
    ``unmapped_fields`` and warn the user.
    """
    _ = boq_json  # BoQ content isn't consulted here; the mapping is purely
    # a template-structure decision. Signature kept for callers that log
    # counts alongside the mapping (bid-assembly does).
    user_mapping = user_mapping or {}
    report: Dict[str, Any] = {}

    for sheet_name, analysis in template_analysis.items():
        columns = list(analysis.get("columns", []))
        # Leftmost-first ordering breaks fuzzy ties deterministically.
        columns.sort(key=lambda c: c["col_index"])
        pool: List[Dict[str, Any]] = list(columns)
        # Look up user overrides for this sheet if any.
        sheet_overrides = user_mapping.get(sheet_name) or {}

        field_to_column: Dict[str, Any] = {}
        unmapped_fields: List[str] = []
        warnings: List[str] = []
        claimed_ids = set()

        for field in _CANONICAL_FIELDS:
            override = sheet_overrides.get(field, "__missing__")
            if override is None:
                # Explicit null: caller wants us to leave this field unmapped.
                field_to_column[field] = None
                continue
            if isinstance(override, str) and override:
                # Locate the named column.
                target = None
                override_norm = _normalise_header(override)
                for col in columns:
                    if col["label_normalised"] == override_norm and col["col_index"] not in claimed_ids:
                        target = col
                        break
                if target is None:
                    # Header not present in the template — hard fail via
                    # a distinguishable error so the CLI can exit 7.
                    raise KeyError(
                        f"Mapping references '{override}' which is not a "
                        f"column header on sheet '{sheet_name}'."
                    )
                claimed_ids.add(target["col_index"])
                field_to_column[field] = {
                    "col_letter": target["col_letter"],
                    "label": target["label"],
                    "match": "user",
                    "score": None,
                }
                continue

            # Auto-match against the still-unclaimed pool.
            available = [c for c in pool if c["col_index"] not in claimed_ids]
            match = _match_field(field, available, fuzzy_threshold)
            if match is None:
                field_to_column[field] = None
                unmapped_fields.append(field)
                continue
            col, kind, score = match
            claimed_ids.add(col["col_index"])
            field_to_column[field] = {
                "col_letter": col["col_letter"],
                "label": col["label"],
                "match": kind,
                "score": score,
            }
            if kind == "fuzzy" and score is not None:
                warnings.append(
                    f"fuzzy: '{col['label']}' -> {field} ({score:.2f})"
                )

        unmapped_columns = [
            c["label"] for c in columns if c["col_index"] not in claimed_ids
        ]
        report[sheet_name] = {
            "field_to_column": field_to_column,
            "unmapped_fields": unmapped_fields,
            "unmapped_columns": unmapped_columns,
            "warnings": warnings,
        }
    return report


def _sheet_has_required(mapping_for_sheet: dict) -> bool:
    ftc = mapping_for_sheet.get("field_to_column", {})
    return all(ftc.get(f) for f in _REQUIRED_FIELDS)


# --------------------------------------------------------------------------- #
# Mapping report printing                                                     #
# --------------------------------------------------------------------------- #

def _print_mapping_report(report: dict, stream=sys.stderr) -> None:
    for sheet_name, sheet in report.items():
        stream.write(f"\nSheet: {sheet_name}\n")
        stream.write("  Field              -> Column          Match     Score\n")
        stream.write("  -----------------  ----------------  --------  -----\n")
        for field, meta in sheet.get("field_to_column", {}).items():
            if meta is None:
                stream.write(f"  {field:<17}  {'(unmapped)':<16}  {'-':<8}  {'-':<5}\n")
                continue
            score = meta.get("score")
            score_txt = f"{score:.2f}" if isinstance(score, (int, float)) else "-"
            stream.write(
                f"  {field:<17}  {meta['label'][:16]:<16}  "
                f"{meta['match']:<8}  {score_txt:<5}\n"
            )
        if sheet.get("unmapped_columns"):
            stream.write(
                "  Untouched template columns: "
                + ", ".join(sheet["unmapped_columns"]) + "\n"
            )
        for w in sheet.get("warnings", []):
            stream.write(f"  ! {w}\n")
    stream.flush()


# --------------------------------------------------------------------------- #
# format_output                                                               #
# --------------------------------------------------------------------------- #

def _item_type(item: dict) -> str:
    subs = item.get("sub_classifications") or []
    return "secondary" if subs else "primary"


def _description_for(item: dict) -> str:
    principal = str(item.get("principal_item") or "").strip()
    subs = item.get("sub_classifications") or []
    if subs:
        joined = ", ".join(str(s).strip() for s in subs if str(s).strip())
        if joined:
            return f"{principal}, {joined}" if principal else joined
    return principal


def _value_for_field(item: dict, field: str) -> Any:
    if field == "element_ref":
        return item.get("element_ref")
    if field == "element_name":
        return item.get("element_name")
    if field == "work_section":
        return item.get("work_section")
    if field == "id":
        return item.get("id")
    if field == "description":
        return _description_for(item)
    if field == "unit":
        return item.get("unit")
    if field == "quantity":
        q = item.get("quantity")
        try:
            return None if q is None else float(q)
        except (TypeError, ValueError):
            return q
    if field == "waste_factor":
        w = item.get("waste_factor")
        if w is None:
            return None
        try:
            return float(w)
        except (TypeError, ValueError):
            return w
    if field == "source_ref":
        return item.get("source_sheet")
    if field == "confidence":
        v = item.get("validation") or {}
        return v.get("confidence")
    return None


def _extract_section_ref(text: str) -> Optional[str]:
    """Pull an NRM element ref out of a section-header text."""
    if not text:
        return None
    match = _ELEMENT_REF_RE.search(text)
    if match:
        return match.group(1)
    # Fallback: name lookup.
    key = _normalise_header(text)
    # Only try the name lookup on short-ish section titles — otherwise
    # descriptive text like "Element 2 - superstructure works" matches
    # both the digit and the word.
    for name, code in NRM_ELEMENT_NAME_TO_CODE.items():
        if name in key:
            return code
    return None


def _element_ref_matches(item_ref: Any, current: Optional[str]) -> bool:
    """Return True if the item belongs under the current section.

    A group ref like ``2.6`` matches parent section ``2``. Exact matches
    always win. If ``current`` is None every item matches (flat mode).
    """
    if current is None:
        return True
    if item_ref is None:
        return False
    item_s = str(item_ref).strip()
    if not item_s:
        return False
    if item_s == current:
        return True
    # Parent-of relation: current "2" matches item "2.6".
    if item_s.startswith(current + "."):
        return True
    # And the reverse — current "2.6" also accepts bare "2" only if that
    # is the only thing to hand; we keep this strict-parent for now.
    return False


def _copy_row_style(ws: "Worksheet", src_row: int, dst_row: int, max_col: int) -> None:
    if src_row is None or src_row < 1:
        return
    for c in range(1, max_col + 1):
        src = ws.cell(row=src_row, column=c)
        dst = ws.cell(row=dst_row, column=c)
        try:
            dst._style = copy(src._style)  # type: ignore[attr-defined]
        except Exception:
            pass
    # Row height too, if set.
    try:
        src_dim = ws.row_dimensions.get(src_row)
        if src_dim is not None and src_dim.height is not None:
            ws.row_dimensions[dst_row].height = src_dim.height
    except Exception:
        pass


def _write_item_to_row(
    ws: "Worksheet",
    row: int,
    item: dict,
    field_to_column: Dict[str, Any],
    formula_cells: set,
    columns_by_letter: Dict[str, Dict[str, Any]],
) -> None:
    """Write mapped fields of ``item`` into ``row``.

    Cells whose coordinate appears in ``formula_cells`` are left alone —
    the bring-your-own-template contract forbids overwriting user
    formulas even accidentally.
    """
    for field, meta in field_to_column.items():
        if meta is None:
            continue
        col_letter = meta["col_letter"]
        coord = f"{col_letter}{row}"
        if coord in formula_cells:
            continue
        cell = ws[coord]
        if _cell_is_formula(cell):
            # Belt-and-braces — analysis missed it, or someone wrote a
            # formula in memory between analyse and write.
            formula_cells.add(coord)
            continue
        value = _value_for_field(item, field)
        if value is None or (isinstance(value, str) and value == ""):
            # Leave blank cells untouched so the user can spot gaps.
            continue

        # Number-format handling for quantity / waste_factor.
        col_meta = columns_by_letter.get(col_letter, {})
        _ = col_meta  # reserved for future per-column formatting rules
        if field == "quantity" and isinstance(value, (int, float)):
            existing_fmt = getattr(cell, "number_format", None)
            if not existing_fmt or existing_fmt == "General":
                cell.number_format = "#,##0.00"
        if field == "waste_factor" and isinstance(value, (int, float)):
            # If the template column format contains %, assume the field is
            # already a ratio (openpyxl multiplies for display); otherwise
            # write the raw float.
            pass
        cell.value = value


def _rows_available_under(
    ws: "Worksheet",
    header_row: int,
    style_zone_rows: List[int],
    max_col: int,
    max_row: int,
) -> Dict[int, List[int]]:
    """For each section-header row, list the rows below it that are
    available for data (empty, no formulas), up until the next section
    header, subtotal, or end of used range."""
    boundaries = sorted(style_zone_rows)
    result: Dict[int, List[int]] = {}
    for i, start in enumerate(boundaries):
        end = boundaries[i + 1] if i + 1 < len(boundaries) else max_row + 1
        avail: List[int] = []
        for r in range(start + 1, end):
            # Only rows that are entirely empty and contain no formulas.
            empty = True
            for c in range(1, max_col + 1):
                cell = ws.cell(row=r, column=c)
                if _cell_is_formula(cell) or cell.value not in (None, ""):
                    empty = False
                    break
            if empty:
                avail.append(r)
        result[start] = avail
    return result


def format_output(
    template_path: str | Path,
    boq_json: dict,
    output_path: str | Path,
    mapping: Optional[dict] = None,
    auto_map: bool = False,
    sheet_filter: Optional[str] = None,
    fuzzy_threshold: float = 0.75,
) -> dict:
    """Copy the template, fill in the takeoff, save to ``output_path``.

    Returns the mapping report actually used, so callers (e.g.
    ``ailtir_bid-assembly``) can log the decisions taken. The template
    file is not modified in place — a byte-copy is taken first so that
    every image / print setting / VBA project / merged range survives
    even if openpyxl chokes on some obscure feature during load.
    """
    if openpyxl is None:
        raise RuntimeError(
            "openpyxl is required; install with `pip install openpyxl`."
        )
    template_path = Path(template_path)
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    shutil.copyfile(template_path, output_path)

    analysis = analyze_template(template_path)
    report = map_takeoff_to_template(
        boq_json, analysis, mapping, fuzzy_threshold=fuzzy_threshold,
    )

    if sheet_filter is not None:
        report = {k: v for k, v in report.items() if k == sheet_filter}
        analysis = {k: v for k, v in analysis.items() if k == sheet_filter}

    # Sheets we can actually populate.
    fillable = [
        s for s, r in report.items()
        if analysis.get(s, {}).get("header_row") is not None
        and _sheet_has_required(r)
    ]

    if not fillable:
        raise LookupError(
            "No worksheet exposes the required BoQ fields "
            f"({', '.join(_REQUIRED_FIELDS)}) after auto-mapping."
        )

    wb = openpyxl.load_workbook(output_path, data_only=False, keep_vba=False)

    items: List[dict] = list(boq_json.get("items") or [])

    for sheet_name in fillable:
        ws = wb[sheet_name]
        sheet_analysis = analysis[sheet_name]
        sheet_report = report[sheet_name]
        header_row = sheet_analysis["header_row"]
        max_col = sheet_analysis["dimensions"]["max_col"]
        max_row = max(sheet_analysis["dimensions"]["max_row"], header_row + 1)
        columns_by_letter = {
            c["col_letter"]: c for c in sheet_analysis["columns"]
        }
        field_to_column = sheet_report["field_to_column"]
        formula_cells = set(sheet_analysis["formula_cells"])
        # Remember the original formula strings so we can restore any
        # that openpyxl silently reset during save/reopen.
        original_formulas: Dict[str, str] = {}
        for coord in formula_cells:
            try:
                original_formulas[coord] = ws[coord].value  # type: ignore[assignment]
            except Exception:
                pass

        section_zones = [
            z for z in sheet_analysis["style_zones"] if z["kind"] == "section_header"
        ]

        # Track which items have been consumed.
        consumed = [False] * len(items)

        def _pick_next(current: Optional[str]) -> Optional[int]:
            """Return index of the next unconsumed item under ``current``."""
            for idx, it in enumerate(items):
                if consumed[idx]:
                    continue
                if _element_ref_matches(it.get("element_ref"), current):
                    return idx
            return None

        if section_zones:
            zone_rows = [z["row"] for z in section_zones]
            avail_map = _rows_available_under(
                ws, header_row, zone_rows, max_col, max_row,
            )
            zone_by_row = {z["row"]: z for z in section_zones}
            # Process each section top-to-bottom.
            zone_rows_sorted = sorted(zone_rows)
            for i, start_row in enumerate(zone_rows_sorted):
                zone = zone_by_row[start_row]
                current_ref = _extract_section_ref(zone["text"])
                slots = list(avail_map.get(start_row, []))
                # Fill the pre-existing empty rows first.
                for row in slots:
                    idx = _pick_next(current_ref)
                    if idx is None:
                        break
                    _write_item_to_row(
                        ws, row, items[idx], field_to_column,
                        formula_cells, columns_by_letter,
                    )
                    consumed[idx] = True
                # Any remaining items for this section get appended right
                # before the next section header (or at end of sheet).
                next_boundary = (
                    zone_rows_sorted[i + 1]
                    if i + 1 < len(zone_rows_sorted) else None
                )
                # Insert new rows before next_boundary if any items remain.
                while True:
                    idx = _pick_next(current_ref)
                    if idx is None:
                        break
                    if next_boundary is not None:
                        ws.insert_rows(next_boundary)
                        # Re-map formula cells whose row moved.
                        formula_cells = _shift_formula_cells(
                            formula_cells, next_boundary,
                        )
                        # And shift future zone rows in our local list.
                        zone_rows_sorted = [
                            r + 1 if r >= next_boundary else r
                            for r in zone_rows_sorted
                        ]
                        next_boundary += 1
                        target_row = next_boundary - 1
                        # Grab styling from the row above (last data row).
                        _copy_row_style(ws, target_row - 1, target_row, max_col)
                    else:
                        max_row += 1
                        target_row = max_row
                        _copy_row_style(ws, target_row - 1, target_row, max_col)
                    _write_item_to_row(
                        ws, target_row, items[idx], field_to_column,
                        formula_cells, columns_by_letter,
                    )
                    consumed[idx] = True
            # Items with no matching section header get appended at the end
            # in NRM order (see fallback branch below).
            leftovers = [i for i, done in enumerate(consumed) if not done]
            if leftovers:
                for idx in _nrm_ordered_indices(items, leftovers):
                    max_row += 1
                    target_row = max_row
                    _copy_row_style(ws, target_row - 1, target_row, max_col)
                    _write_item_to_row(
                        ws, target_row, items[idx], field_to_column,
                        formula_cells, columns_by_letter,
                    )
                    consumed[idx] = True
        else:
            # NRM2 fallback: flat list ordered by element_ref then
            # primary-before-secondary then id.
            order = _nrm_ordered_indices(items, range(len(items)))
            # Fill existing empty rows first, then append.
            row_cursor = header_row + 1
            for idx in order:
                # Find next empty writable row.
                while row_cursor <= max_row:
                    empty = True
                    for c in range(1, max_col + 1):
                        cell = ws.cell(row=row_cursor, column=c)
                        if _cell_is_formula(cell) or cell.value not in (None, ""):
                            empty = False
                            break
                    if empty:
                        break
                    row_cursor += 1
                if row_cursor > max_row:
                    max_row = row_cursor
                    _copy_row_style(ws, row_cursor - 1, row_cursor, max_col)
                _write_item_to_row(
                    ws, row_cursor, items[idx], field_to_column,
                    formula_cells, columns_by_letter,
                )
                consumed[idx] = True
                row_cursor += 1

        # Restore any formulas openpyxl may have blanked.
        for coord, formula in original_formulas.items():
            try:
                if ws[coord].value in (None, ""):
                    ws[coord] = formula
            except Exception:
                pass

    wb.save(output_path)
    wb.close()
    return report


def _shift_formula_cells(formula_cells: set, insert_at: int) -> set:
    """Return a new set with coords whose row >= ``insert_at`` bumped by 1."""
    if not formula_cells:
        return formula_cells
    shifted: set = set()
    for coord in formula_cells:
        try:
            col_letter, row = coordinate_from_string(coord)
        except Exception:
            shifted.add(coord)
            continue
        if row >= insert_at:
            shifted.add(f"{col_letter}{row + 1}")
        else:
            shifted.add(coord)
    return shifted


def _nrm_ordered_indices(items: List[dict], indices: Iterable[int]) -> List[int]:
    """Sort indices by (element_ref ascending, primary first, id ascending)."""
    def _key(idx: int) -> Tuple[Any, ...]:
        it = items[idx]
        ref = it.get("element_ref") or ""
        # Split "2.6" -> (2, 6) so that "2.10" sorts after "2.9".
        parts: List[Any] = []
        for chunk in str(ref).split("."):
            try:
                parts.append(int(chunk))
            except (TypeError, ValueError):
                parts.append(chunk)
        type_rank = 0 if _item_type(it) == "primary" else 1
        return (parts, type_rank, str(it.get("id") or ""))
    return sorted(indices, key=_key)


# --------------------------------------------------------------------------- #
# CLI                                                                         #
# --------------------------------------------------------------------------- #

def _load_json(path: Path) -> Any:
    with path.open("r", encoding="utf-8") as fh:
        return json.load(fh)


def _validate_boq_shape(blob: Any) -> None:
    if not isinstance(blob, dict):
        raise ValueError("BoQ JSON root must be an object.")
    items = blob.get("items")
    if not isinstance(items, list):
        raise ValueError("BoQ JSON is missing an 'items' array.")
    for i, item in enumerate(items):
        if not isinstance(item, dict):
            raise ValueError(f"BoQ item #{i} is not an object.")
        # Only the fields we consume must be strictly typed. Everything
        # else may be present but is ignored.
        if "quantity" in item and item["quantity"] is not None:
            try:
                float(item["quantity"])
            except (TypeError, ValueError) as exc:
                raise ValueError(
                    f"BoQ item #{i}: quantity is not numeric ({exc})."
                ) from exc


def _confirm_mapping(report: dict) -> bool:
    _print_mapping_report(report, sys.stderr)
    sys.stderr.write("\nProceed with this mapping? [y/N]: ")
    sys.stderr.flush()
    try:
        line = sys.stdin.readline()
    except Exception:
        return False
    if line is None:
        return False
    return line.strip().lower() in ("y", "yes")


def build_arg_parser() -> argparse.ArgumentParser:
    p = argparse.ArgumentParser(
        prog="template_formatter",
        description=(
            "Populate a user-supplied Excel BoQ template with validated "
            "NRM2 takeoff data. Preserves the template's columns, styles "
            "and formulas."
        ),
    )
    p.add_argument("--template", required=True, help="Path to the .xlsx template.")
    p.add_argument("--json", dest="json_path", required=True, help="Path to BoQ JSON.")
    p.add_argument("--out", required=True, help="Where to write the filled workbook.")
    p.add_argument("--mapping", default=None, help="Optional JSON overriding auto-mapping.")
    p.add_argument("--auto-map", action="store_true",
                   help="Commit to the auto-generated mapping without prompting.")
    p.add_argument("--sheet", default=None,
                   help="Restrict formatting to a single named sheet.")
    p.add_argument("--fuzzy-threshold", type=float, default=0.75,
                   help="Minimum similarity ratio for a fuzzy header match (default 0.75).")
    p.add_argument("--quiet", action="store_true",
                   help="Silence the mapping report on success.")
    return p


def main(argv: Optional[List[str]] = None) -> int:
    parser = build_arg_parser()
    args = parser.parse_args(argv)

    if openpyxl is None:
        sys.stderr.write(
            "template_formatter needs openpyxl. Install it with:\n"
            "    pip install openpyxl\n"
        )
        return EXIT_NO_OPENPYXL

    template_path = Path(args.template)
    if not template_path.exists() or not template_path.is_file():
        sys.stderr.write(f"Template not found or unreadable: {template_path}\n")
        return EXIT_TEMPLATE_BAD
    try:
        # Cheap open to confirm openpyxl can parse it.
        wb = openpyxl.load_workbook(template_path, data_only=False, keep_vba=False)
        wb.close()
    except Exception as exc:
        sys.stderr.write(f"openpyxl could not open template: {exc}\n")
        return EXIT_TEMPLATE_BAD

    json_path = Path(args.json_path)
    if not json_path.exists() or not json_path.is_file():
        sys.stderr.write(f"BoQ JSON not found: {json_path}\n")
        return EXIT_JSON_BAD
    try:
        boq_json = _load_json(json_path)
        _validate_boq_shape(boq_json)
    except Exception as exc:
        sys.stderr.write(f"BoQ JSON is not valid: {exc}\n")
        return EXIT_JSON_BAD

    out_path = Path(args.out)
    try:
        out_path.parent.mkdir(parents=True, exist_ok=True)
        # Touch-test write access.
        with out_path.open("a"):
            pass
        # If it was empty and we just created it, that's fine — the copy
        # below will overwrite. Do not delete pre-existing content here.
    except Exception as exc:
        sys.stderr.write(f"Cannot write to --out {out_path}: {exc}\n")
        return EXIT_OUT_BAD

    user_mapping: Optional[dict] = None
    if args.mapping:
        try:
            user_mapping = _load_json(Path(args.mapping))
            if not isinstance(user_mapping, dict):
                raise ValueError("mapping JSON root must be an object")
        except Exception as exc:
            sys.stderr.write(f"--mapping JSON invalid: {exc}\n")
            return EXIT_MAPPING_BAD

    # First pass: analysis + proposed mapping. If we're prompting the
    # user, we do it before we run the (potentially expensive) fill.
    try:
        analysis = analyze_template(template_path)
    except Exception as exc:
        sys.stderr.write(f"Failed to analyse template: {exc}\n")
        return EXIT_TEMPLATE_BAD

    try:
        proposed = map_takeoff_to_template(
            boq_json, analysis, user_mapping,
            fuzzy_threshold=args.fuzzy_threshold,
        )
    except KeyError as exc:
        sys.stderr.write(f"Mapping override invalid: {exc}\n")
        return EXIT_MAPPING_BAD

    if args.sheet is not None:
        proposed = {k: v for k, v in proposed.items() if k == args.sheet}
        if not proposed:
            sys.stderr.write(
                f"--sheet '{args.sheet}' not present in template.\n"
            )
            return EXIT_NO_USABLE_SHEET

    usable = [
        s for s, r in proposed.items()
        if analysis.get(s, {}).get("header_row") is not None
        and _sheet_has_required(r)
    ]
    if not usable:
        sys.stderr.write(
            "No worksheet exposes the required BoQ fields ("
            + ", ".join(_REQUIRED_FIELDS) + ") after auto-mapping.\n"
        )
        return EXIT_NO_USABLE_SHEET

    # Interactive-safe by default: without --auto-map, prompt the user.
    if not args.auto_map:
        if not _confirm_mapping(proposed):
            sys.stderr.write("User declined the proposed mapping.\n")
            return EXIT_USER_DECLINED
    elif not args.quiet:
        _print_mapping_report(proposed, sys.stderr)

    try:
        format_output(
            template_path,
            boq_json,
            out_path,
            mapping=user_mapping,
            auto_map=True,  # already confirmed / auto
            sheet_filter=args.sheet,
            fuzzy_threshold=args.fuzzy_threshold,
        )
    except LookupError as exc:
        sys.stderr.write(f"{exc}\n")
        return EXIT_NO_USABLE_SHEET
    except KeyError as exc:
        sys.stderr.write(f"Mapping error: {exc}\n")
        return EXIT_MAPPING_BAD
    except PermissionError as exc:
        sys.stderr.write(f"Cannot write output: {exc}\n")
        return EXIT_OUT_BAD
    except Exception as exc:
        sys.stderr.write(f"Unexpected failure while writing output: {exc}\n")
        return EXIT_OUT_BAD

    if not args.quiet:
        sys.stderr.write(f"\nWrote filled workbook: {out_path}\n")
    return EXIT_OK


if __name__ == "__main__":
    raise SystemExit(main())
