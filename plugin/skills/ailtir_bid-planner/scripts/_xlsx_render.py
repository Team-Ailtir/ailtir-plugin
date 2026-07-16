"""Deterministic Ailtir workbook renderer.

The script owns structure (tab titles, order, headers) and styling; the model
supplies row content via a --data JSON blob. Scripts define a CORE_TABS
skeleton and pass model data through merge_rows() into build_workbook().

Bundled verbatim into each skill's scripts/ directory because Cowork does not
support reliable cross-skill Python imports (scripts run with cwd = session
root, and are invoked by absolute path).
"""
from __future__ import annotations

import json
import sys
from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("openpyxl not installed. Run: python3 -m pip install openpyxl", file=sys.stderr)
    raise

# Ailtir brand palette
NAVY = "0A1128"
PURPLE = "7C3AED"
LIGHT = "F5F7FA"
WHITE = "FFFFFF"
AMBER = "F59E0B"
NA_GREY = "9CA3AF"

HEADER_FONT = Font(bold=True, color=WHITE, size=11, name="Space Grotesk")
HEADER_FILL = PatternFill("solid", fgColor=NAVY)
SECTION_FONT = Font(bold=True, color=WHITE, size=11, name="Space Grotesk")
SECTION_FILL = PatternFill("solid", fgColor=PURPLE)
TITLE_FONT = Font(bold=True, color=WHITE, size=14, name="Space Grotesk")
BODY_FONT = Font(size=10, name="Inter")
LABEL_FONT = Font(bold=True, size=10, name="Inter")
NA_FONT = Font(italic=True, color=NA_GREY, size=10, name="Inter")
BANNER_FONT = Font(italic=True, color=PURPLE, size=10, name="Inter")
_THIN = Side(style="thin", color="D9D9D9")
BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="top", wrap_text=True)


def load_data(path):
    if not path:
        return {}
    return json.loads(Path(path).read_text(encoding="utf-8"))


def _style_header(ws, row, ncols, fill=HEADER_FILL, font=HEADER_FONT):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = font
        cell.fill = fill
        cell.alignment = CENTER
        cell.border = BORDER


def _autosize(ws, ncols, widths=None):
    for c in range(1, ncols + 1):
        w = widths[c - 1] if widths and c - 1 < len(widths) else 22
        ws.column_dimensions[get_column_letter(c)].width = w


def _render_cover(ws, cover):
    ws.title = cover.get("sheet_title", "1. Bid Summary")
    ws.merge_cells("A1:F1")
    ws["A1"] = cover.get("title", "AILTIR BID PLAN")
    ws["A1"].font = TITLE_FONT
    ws["A1"].fill = HEADER_FILL
    ws["A1"].alignment = CENTER
    row = 3
    for label, value in cover.get("fields", []):
        ws.cell(row=row, column=1, value=label).font = LABEL_FONT
        ws.cell(row=row, column=2, value=value).font = BODY_FONT
        row += 1
    _autosize(ws, 2, [26, 60])


def _render_grid(ws, start_row, headers, rows, na_note):
    r = start_row
    if headers:
        for i, h in enumerate(headers, 1):
            ws.cell(row=r, column=i, value=h)
        _style_header(ws, r, len(headers))
        r += 1
    if rows:
        for row_vals in rows:
            for i, val in enumerate(row_vals, 1):
                cell = ws.cell(row=r, column=i, value=val)
                cell.font = BODY_FONT
                cell.alignment = LEFT
                cell.border = BORDER
            r += 1
    elif na_note:
        ws.cell(row=r, column=1, value=na_note).font = NA_FONT
        r += 1
    _autosize(ws, max(len(headers) if headers else 1, 1))
    return r


def _render_tab(wb, spec):
    ws = wb.create_sheet(spec["title"])
    r = 1
    sections = spec.get("sections")
    if sections:
        for sec in sections:
            ncols = max(len(sec.get("headers", [])), 1)
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=ncols)
            ws.cell(row=r, column=1, value=sec["heading"])
            _style_header(ws, r, ncols, fill=SECTION_FILL, font=SECTION_FONT)
            r += 1
            r = _render_grid(ws, r, sec.get("headers", []), sec.get("rows", []), sec.get("na_note"))
            r += 1
    else:
        r = _render_grid(ws, r, spec.get("headers", []), spec.get("rows", []), spec.get("na_note"))
    if spec.get("banner"):
        r += 1
        ws.cell(row=r, column=1, value=spec["banner"]).font = BANNER_FONT
    return ws


def build_workbook(cover, tabs):
    wb = Workbook()
    _render_cover(wb.active, cover)
    for spec in tabs:
        _render_tab(wb, spec)
    return wb


def merge_rows(core_tabs, data):
    filled = []
    supplied = data.get("tabs", {})
    for spec in core_tabs:
        out = dict(spec)
        d = supplied.get(spec.get("key"), {})
        if "sections" in spec:
            secs = []
            sd_all = d.get("sections", {})
            for base in spec["sections"]:
                ms = dict(base)
                sd = sd_all.get(base.get("key"), {})
                ms["rows"] = sd.get("rows", [])
                if "na_note" in sd:
                    ms["na_note"] = sd["na_note"]
                secs.append(ms)
            out["sections"] = secs
        else:
            out["rows"] = d.get("rows", [])
            if "na_note" in d:
                out["na_note"] = d["na_note"]
        filled.append(out)
    for opt in data.get("optional_tabs", []):
        filled.append(opt)
    return filled
