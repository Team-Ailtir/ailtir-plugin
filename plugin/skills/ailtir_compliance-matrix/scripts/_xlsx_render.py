"""Ailtir workbook renderer — permissive edition.

The engine owns: tab existence and ordering, Ailtir brand styling, Excel-safe
sheet titles, and output file I/O. The model owns: every header, every data
row, section structure, callout text, and optional extra tabs.

Scripts define a CORE_TABS list of {key, title} pairs that guarantee certain
tabs are always present. The model supplies all content for each tab via a
--data JSON blob. See DATA CONTRACT in each skill's SKILL.md.

Bundled verbatim into each skill's scripts/ because Cowork does not support
reliable cross-skill Python imports.
"""
from __future__ import annotations

import json
import re
import sys
from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("openpyxl not installed. Run: python3 -m pip install openpyxl", file=sys.stderr)
    raise

# Ailtir brand palette
NAVY    = "0A1128"
PURPLE  = "7C3AED"
WHITE   = "FFFFFF"
AMBER   = "F59E0B"
NA_GREY = "9CA3AF"

HEADER_FONT  = Font(bold=True, color=WHITE,   size=11, name="Space Grotesk")
HEADER_FILL  = PatternFill("solid", fgColor=NAVY)
SECTION_FONT = Font(bold=True, color=WHITE,   size=11, name="Space Grotesk")
SECTION_FILL = PatternFill("solid", fgColor=PURPLE)
TITLE_FONT   = Font(bold=True, color=WHITE,   size=14, name="Space Grotesk")
BODY_FONT    = Font(size=10, name="Inter")
LABEL_FONT   = Font(bold=True, size=10, name="Inter")
NA_FONT      = Font(italic=True, color=NA_GREY, size=10, name="Inter")
BANNER_FONT  = Font(italic=True, color=PURPLE,  size=10, name="Inter")
CALLOUT_FONT = Font(bold=True, color=NAVY, size=12, name="Space Grotesk")
CALLOUT_FILL = PatternFill("solid", fgColor=AMBER)
_THIN  = Side(style="thin", color="D9D9D9")
BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT   = Alignment(horizontal="left",   vertical="top",    wrap_text=True)


def load_data(path):
    if not path:
        return {}
    try:
        return json.loads(Path(path).read_text(encoding="utf-8"))
    except FileNotFoundError:
        print(f"--data file not found: {path}", file=sys.stderr)
        raise SystemExit(1)
    except json.JSONDecodeError as e:
        print(f"--data is not valid JSON ({path}): {e}", file=sys.stderr)
        raise SystemExit(1)


def save_workbook(wb, output):
    out = Path(output)
    if out.parent and not out.parent.exists():
        out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output)


_INVALID_TITLE = re.compile(r"[\\*?:/\[\]]")

def safe_sheet_title(title):
    return _INVALID_TITLE.sub("-", title)[:31]


def _style_header(ws, row, ncols, fill=HEADER_FILL, font=HEADER_FONT):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = font
        cell.fill = fill
        cell.alignment = CENTER
        cell.border = BORDER


def _autosize(ws, ncols, widths=None):
    for c in range(1, ncols + 1):
        w = widths[c - 1] if widths and c - 1 < len(widths) else 22
        ws.column_dimensions[get_column_letter(c)].width = w


def _render_cover(ws, cover):
    ws.title = safe_sheet_title(cover.get("sheet_title", "1. Bid Summary"))
    ws.merge_cells("A1:F1")
    ws["A1"] = cover.get("title", "AILTIR BID PLAN")
    ws["A1"].font = TITLE_FONT
    ws["A1"].fill = HEADER_FILL
    ws["A1"].alignment = CENTER
    row = 3
    for label, value in cover.get("fields", []):
        ws.cell(row=row, column=1, value=label).font = LABEL_FONT
        ws.cell(row=row, column=2, value=value).font = BODY_FONT
        row += 1
    _autosize(ws, 2, [26, 60])


def _render_grid(ws, start_row, headers, rows, na_note=None, widths=None):
    r = start_row
    ncols = max(len(headers), 1)
    if headers:
        for i, h in enumerate(headers, 1):
            ws.cell(row=r, column=i, value=h)
        _style_header(ws, r, ncols)
        r += 1
    if rows:
        for row_vals in rows:
            for i, val in enumerate(row_vals, 1):
                cell = ws.cell(row=r, column=i, value=val)
                cell.font = BODY_FONT
                cell.alignment = LEFT
                cell.border = BORDER
            r += 1
    elif na_note:
        ws.cell(row=r, column=1, value=na_note).font = NA_FONT
        r += 1
    _autosize(ws, ncols, widths)
    return r


def _render_tab(wb, spec):
    ws = wb.create_sheet(safe_sheet_title(spec["title"]))
    r = 1
    sections = spec.get("sections")
    ncols = max(
        max((len(s.get("headers", [])) for s in sections), default=1)
        if sections else len(spec.get("headers", [])),
        1,
    )
    if spec.get("callout"):
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=ncols)
        for c in range(1, ncols + 1):
            cell = ws.cell(row=r, column=c)
            cell.font = CALLOUT_FONT
            cell.fill = CALLOUT_FILL
            cell.alignment = CENTER
        ws.cell(row=r, column=1, value=spec["callout"])
        r += 2
    if sections:
        for sec in sections:
            sec_ncols = max(len(sec.get("headers", [])), 1)
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=sec_ncols)
            ws.cell(row=r, column=1, value=sec.get("heading", ""))
            _style_header(ws, r, sec_ncols, fill=SECTION_FILL, font=SECTION_FONT)
            r += 1
            r = _render_grid(ws, r, sec.get("headers", []), sec.get("rows", []),
                             sec.get("na_note"), sec.get("widths"))
            r += 1
    else:
        r = _render_grid(ws, r, spec.get("headers", []), spec.get("rows", []),
                         spec.get("na_note"), spec.get("widths"))
    if spec.get("banner"):
        r += 1
        ws.cell(row=r, column=1, value=spec["banner"]).font = BANNER_FONT
    return ws


def build_workbook(cover, tabs):
    wb = Workbook()
    _render_cover(wb.active, cover)
    for spec in tabs:
        _render_tab(wb, spec)
    return wb


def merge_rows(core_tabs, data):
    """Merge CORE_TABS skeleton with model-supplied data.

    CORE_TABS entries need only {key, title}. The model supplies headers, rows,
    sections, callout, and banner via data["tabs"][key]. Tabs with no matching
    data entry are rendered empty (header row only if the spec provides headers).
    """
    filled = []
    supplied = data.get("tabs", {})
    for spec in core_tabs:
        d = supplied.get(spec.get("key"), {})
        out = {"title": spec["title"]}
        for k in ("callout", "banner", "na_note"):
            v = d.get(k) or spec.get(k)
            if v:
                out[k] = v
        if "sections" in d:
            clean_secs = []
            for s in d["sections"]:
                sec = {
                    "heading": s.get("heading", ""),
                    "headers": s.get("headers", []),
                    "rows":    s.get("rows", []),
                }
                for k in ("na_note", "widths"):
                    if k in s:
                        sec[k] = s[k]
                clean_secs.append(sec)
            out["sections"] = clean_secs
        else:
            out["headers"] = d.get("headers", spec.get("headers", []))
            out["rows"]    = d.get("rows", [])
            for k in ("na_note", "widths"):
                if k in d:
                    out[k] = d[k]
        filled.append(out)
    for opt in data.get("optional_tabs", []):
        filled.append(opt)
    return filled
